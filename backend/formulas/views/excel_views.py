import json
import logging
from copy import copy
from io import BytesIO
from django.db import models
from django.shortcuts import get_object_or_404
from django.http import HttpResponse, JsonResponse
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.cell.cell import MergedCell
from rest_framework import viewsets, status
from rest_framework.decorators import api_view, permission_classes, action
from rest_framework.permissions import AllowAny
from rest_framework.response import Response
from ..models import ExcelTemplate, Protocol, Calculation
from ..serializers import ExcelTemplateSerializer
from ..utils.excel_utils import (
    A4_HEIGHT_POINTS,
    DEFAULT_ROW_HEIGHT,
    save_row_dimensions,
    restore_row_dimensions,
    format_result,
    get_content_height,
    find_table_headers,
    get_table_header_rows,
)

logger = logging.getLogger(__name__)


@permission_classes([AllowAny])
class ExcelTemplateViewSet(viewsets.ModelViewSet):
    serializer_class = ExcelTemplateSerializer
    queryset = ExcelTemplate.objects.all()

    def get_queryset(self):
        queryset = ExcelTemplate.objects.all()

        # Получаем параметры из запроса
        laboratory = self.request.query_params.get("laboratory")
        department = self.request.query_params.get("department")

        # Фильтруем по лаборатории
        if laboratory:
            queryset = queryset.filter(laboratory=laboratory)

            # Если указано подразделение, добавляем его в фильтр
            if department:
                queryset = queryset.filter(
                    models.Q(department=department)  # Шаблоны конкретного подразделения
                    | models.Q(department__isnull=True)  # И общие шаблоны лаборатории
                )

        return queryset

    def create(self, request, *args, **kwargs):
        name = request.data.get("name")
        laboratory = request.data.get("laboratory")

        if not laboratory:
            return Response(
                {"error": "Необходимо указать лабораторию"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        file = request.FILES.get("file")
        if not file:
            return Response(
                {"error": "Файл не предоставлен"}, status=status.HTTP_400_BAD_REQUEST
            )

        file_content = file.read()

        data = {
            "name": name,
            "version": "v1",
            "file_name": file.name,
            "is_active": True,
            "laboratory": laboratory,
            "department": request.data.get("department"),
        }

        serializer = self.get_serializer(data=data)
        serializer.is_valid(raise_exception=True)

        # Проверяем существование активного шаблона
        if ExcelTemplate.objects.filter(
            name=name,
            laboratory=laboratory,
            department=request.data.get("department"),
            is_active=True,
        ).exists():
            return Response(
                {
                    "error": "Активный шаблон с таким именем уже существует для данной лаборатории и подразделения"
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        template = serializer.save(file=file_content)

        return Response(serializer.data, status=status.HTTP_201_CREATED)

    def retrieve(self, request, *args, **kwargs):
        instance = self.get_object()
        logger.info(
            f"Запрос шаблона: id={instance.id}, name={instance.name}, version={instance.version}"
        )
        logger.info(f"Параметры запроса: {request.query_params}")

        if request.query_params.get("download") == "true":
            logger.info(f"Скачивание файла: {instance.file_name}")

            response = HttpResponse(
                instance.file,
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            response["Content-Disposition"] = (
                f'attachment; filename="{instance.file_name}"'
            )
            return response
        return super().retrieve(request, *args, **kwargs)

    @action(detail=True, methods=["post"])
    def update_template(self, request, pk=None):
        """
        Обновляет существующий шаблон, создавая новую версию
        """
        current_template = self.get_object()

        next_version = current_template.deactivate()

        new_template = ExcelTemplate.objects.create(
            name=current_template.name,
            version=next_version,
            file=request.data.get("file", current_template.file),
            file_name=current_template.file_name,
            is_active=True,
            laboratory=current_template.laboratory,
            department=current_template.department,
        )

        serializer = self.get_serializer(new_template)
        return Response(serializer.data)

    @action(detail=True, methods=["patch"])
    def update_selection_conditions(self, request, pk=None):
        """
        Обновляет условия отбора для шаблона
        """
        template = self.get_object()

        selection_conditions = request.data.get("selection_conditions")
        if not isinstance(selection_conditions, list):
            return Response(
                {"error": "Условия отбора должны быть списком"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Проверяем, что каждый элемент содержит только name и unit
        for condition in selection_conditions:
            if not isinstance(condition, dict):
                return Response(
                    {"error": "Каждое условие должно быть объектом"},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            if not all(key in condition for key in ["name", "unit"]):
                return Response(
                    {"error": "Каждое условие должно содержать поля 'name' и 'unit'"},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            if len(condition.keys()) > 2:
                return Response(
                    {
                        "error": "Каждое условие должно содержать только поля 'name' и 'unit'"
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )

        # Создаем новую версию шаблона
        next_version = template.deactivate()

        new_template = ExcelTemplate.objects.create(
            name=template.name,
            version=next_version,
            file=template.file,
            file_name=template.file_name,
            is_active=True,
            laboratory=template.laboratory,
            department=template.department,
            selection_conditions=selection_conditions,
        )

        serializer = self.get_serializer(new_template)
        return Response(serializer.data)

    @action(detail=True, methods=["patch"])
    def update_accreditation_header_row(self, request, pk=None):
        """
        Обновляет номер строки шапки аккредитации для шаблона
        """
        template = self.get_object()

        accreditation_header_row = request.data.get("accreditation_header_row")
        if (
            not isinstance(accreditation_header_row, int)
            or accreditation_header_row < 1
        ):
            return Response(
                {"error": "Номер строки должен быть положительным целым числом"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Обновляем существующий шаблон без создания новой версии
        template.accreditation_header_row = accreditation_header_row
        template.save()

        serializer = self.get_serializer(template)
        return Response(serializer.data)

    @action(detail=False, methods=["get"], url_path="by-laboratory")
    def by_laboratory(self, request):
        """
        Получение шаблонов по лаборатории и подразделению.
        Если подразделение указано, возвращает шаблоны этого подразделения и общие шаблоны лаборатории.
        Если подразделение не указано, возвращает только общие шаблоны лаборатории.
        """
        laboratory = request.query_params.get("laboratory")
        department = request.query_params.get("department")

        if not laboratory:
            return Response(
                {"error": "Необходимо указать лабораторию"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        queryset = self.get_queryset().filter(laboratory=laboratory)

        if department:
            queryset = queryset.filter(
                models.Q(department=department) | models.Q(department__isnull=True)
            )
        else:
            queryset = queryset.filter(department__isnull=True)

        queryset = queryset.order_by("name", "-version")

        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)


@api_view(["POST"])
@permission_classes([AllowAny])
def save_excel(request):
    """
    Сохранение изменений шаблона протокола
    """
    try:
        data = json.loads(request.POST.get("data", "[]"))
        styles = json.loads(request.POST.get("styles", "{}"))
        template_id = request.POST.get("template_id")
        section = request.POST.get("section")

        if not template_id:
            return Response({"error": "Не указан ID шаблона"}, status=400)

        # Получаем текущий шаблон
        current_template = ExcelTemplate.objects.get(id=template_id)

        # Деактивируем текущий шаблон и получаем следующую версию
        next_version = current_template.deactivate()

        file_stream = BytesIO(current_template.file)
        workbook = load_workbook(file_stream, data_only=False)
        worksheet = workbook.active

        # В зависимости от типа протокола применяем разную логику
        if section == "header":
            # Находим существующие метки в файле
            start_header_row = None
            end_header_row = None
            for row_idx in range(1, worksheet.max_row + 1):
                cell_value = worksheet.cell(row=row_idx, column=1).value
                if cell_value == "{{start_header}}":
                    start_header_row = row_idx
                elif cell_value == "{{end_header}}":
                    end_header_row = row_idx

            # Проверяем наличие меток
            if start_header_row is None or end_header_row is None:
                error_message = "В файле не найдены метки {{start_header}} и {{end_header}}. Добавьте метки в шаблон для редактирования шапки."
                logger.error(error_message)
                return Response(
                    {"error": error_message},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            # Если нужно больше места между метками
            if end_header_row - start_header_row - 1 < len(data):
                # Сдвигаем данные после end_header вниз
                shift = len(data) - (end_header_row - start_header_row - 1)
                for row_idx in range(worksheet.max_row, end_header_row - 1, -1):
                    for col_idx in range(1, worksheet.max_column + 1):
                        source_cell = worksheet.cell(row=row_idx, column=col_idx)
                        target_cell = worksheet.cell(
                            row=row_idx + shift, column=col_idx
                        )

                        # Если исходная ячейка объединена, находим основную ячейку
                        if isinstance(source_cell, MergedCell):
                            for merge_range in worksheet.merged_cells.ranges:
                                min_row = merge_range.min_row
                                max_row = merge_range.max_row
                                min_col = merge_range.min_col
                                max_col = merge_range.max_col

                                if (
                                    min_row <= row_idx <= max_row
                                    and min_col <= col_idx <= max_col
                                ):
                                    source_cell = worksheet.cell(
                                        row=min_row, column=min_col
                                    )
                                    # Создаем новый диапазон объединения со сдвигом
                                    worksheet.merge_cells(
                                        start_row=min_row + shift,
                                        start_column=min_col,
                                        end_row=max_row + shift,
                                        end_column=max_col,
                                    )
                                    break

                        # Копируем значение и стили
                        if not isinstance(target_cell, MergedCell):
                            target_cell.value = source_cell.value
                            if source_cell.has_style:
                                target_cell._style = copy(source_cell._style)
                end_header_row += shift

            # Сохраняем информацию об объединенных ячейках
            merged_ranges = []
            for merge_range in worksheet.merged_cells.ranges:
                if start_header_row < merge_range.min_row < end_header_row:
                    merged_ranges.append(
                        {
                            "min_row": merge_range.min_row,
                            "max_row": merge_range.max_row,
                            "min_col": merge_range.min_col,
                            "max_col": merge_range.max_col,
                        }
                    )

            # Очищаем старые данные между метками и разъединяем ячейки
            for row_idx in range(start_header_row + 1, end_header_row):
                # Находим и удаляем объединения ячеек в этой области
                ranges_to_remove = []
                for merge_range in worksheet.merged_cells.ranges:
                    min_row = merge_range.min_row
                    max_row = merge_range.max_row

                    if min_row <= row_idx <= max_row:
                        ranges_to_remove.append(merge_range)

                for merge_range in ranges_to_remove:
                    worksheet.unmerge_cells(
                        start_row=merge_range.min_row,
                        start_column=merge_range.min_col,
                        end_row=merge_range.max_row,
                        end_column=merge_range.max_col,
                    )

                worksheet.cell(row=row_idx, column=1).value = None

            # Записываем новые данные
            for idx, row_data in enumerate(data):
                value = str(row_data[0]) if row_data[0] is not None else ""
                target_row = start_header_row + 1 + idx
                target_col = 1
                cell = worksheet.cell(row=target_row, column=target_col)
                cell.value = value

                # Применяем стили, если они есть для данной ячейки
                cell_key = f"{idx}-0"
                if styles and cell_key in styles:
                    style = styles[cell_key]

                    font_size = style.get("fontSize", "14")
                    if isinstance(font_size, str):
                        font_size = font_size.replace("px", "")
                        try:
                            font_size = int(float(font_size))
                        except (ValueError, TypeError):
                            font_size = 14

                    font = Font(
                        name="Times New Roman",
                        bold=style.get("fontWeight") == "bold",
                        italic=style.get("fontStyle") == "italic",
                        size=font_size,
                    )
                    cell.font = font

                    horizontal_align = "center"
                    if style.get("textAlign") == "left":
                        horizontal_align = "left"
                    elif style.get("textAlign") == "right":
                        horizontal_align = "right"

                    alignment = Alignment(
                        horizontal=horizontal_align,
                        vertical="center",
                        wrap_text=True,
                    )
                    cell.alignment = alignment

            # Восстанавливаем объединенные ячейки
            for merge_info in merged_ranges:
                worksheet.merge_cells(
                    start_row=merge_info["min_row"],
                    start_column=merge_info["min_col"],
                    end_row=merge_info["max_row"],
                    end_column=merge_info["max_col"],
                )

        else:
            logger.error(f"Неизвестная секция для редактирования: {section}")
            return Response(
                {"error": f"Неизвестная секция: {section}"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        output = BytesIO()
        workbook.save(output)
        file_content = output.getvalue()

        # Создаем новый шаблон с измененным содержимым
        new_template = ExcelTemplate.objects.create(
            name=current_template.name,
            version=next_version,
            file=file_content,
            file_name=current_template.file_name,
            is_active=True,
            laboratory=current_template.laboratory,
            department=current_template.department,
        )

        return Response(
            {
                "message": "Файл успешно обновлен",
                "version": next_version,
                "template_id": new_template.id,
            },
            status=status.HTTP_200_OK,
        )

    except Exception as e:
        logger.error(f"Ошибка при сохранении Excel файла: {str(e)}", exc_info=True)
        return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(["GET"])
@permission_classes([AllowAny])
def get_excel_styles(request):
    """
    Получает стили для ячеек в файле
    """
    template_id = request.query_params.get("template_id")
    section = request.query_params.get("section")

    logger.info(f"Запрос стилей Excel: template_id={template_id}, section={section}")

    if not template_id:
        return Response({"error": "Не указан ID шаблона"}, status=400)

    try:
        template = ExcelTemplate.objects.get(id=template_id)
        logger.info(
            f"Найден шаблон: id={template.id}, name={template.name}, version={template.version}"
        )

        file_stream = BytesIO(template.file)
        workbook = load_workbook(file_stream, data_only=False)
        worksheet = workbook.active

        styles = {}

        # Ищем метки в файле
        start_header_row = None
        end_header_row = None
        for row_idx in range(1, worksheet.max_row + 1):
            cell_value = worksheet.cell(row=row_idx, column=1).value
            if cell_value == "{{start_header}}":
                start_header_row = row_idx
            elif cell_value == "{{end_header}}":
                end_header_row = row_idx
                break

        # Проверяем наличие меток
        if start_header_row is None or end_header_row is None:
            error_message = "В файле не найдены метки {{start_header}} и {{end_header}}. Добавьте метки в шаблон для редактирования шапки."
            logger.error(error_message)
            return Response(
                {"error": error_message},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Получаем стили для ячеек между метками
        for row_idx in range(start_header_row + 1, end_header_row):
            cell = worksheet.cell(row=row_idx, column=1)
            cell_key = f"{row_idx - start_header_row - 1}-0"

            if cell.font:
                style = {}
                if cell.font.bold:
                    style["fontWeight"] = "bold"
                if cell.font.italic:
                    style["fontStyle"] = "italic"
                if cell.font.size:
                    style["fontSize"] = f"{cell.font.size}px"

                if style:
                    styles[cell_key] = style

        return Response({"styles": styles}, status=status.HTTP_200_OK)
    except Exception as e:
        logger.error(f"Ошибка при получении стилей Excel: {str(e)}", exc_info=True)
        return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
