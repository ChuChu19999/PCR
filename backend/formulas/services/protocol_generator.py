import logging
from io import BytesIO
import openpyxl
from copy import copy
from django.http import HttpResponse, HttpResponseBadRequest
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import AllowAny
from ..models import Protocol, ExcelTemplate, ResearchObjectMethod
from django.db import models
from ..models import Equipment
from django.utils import formats
from ..utils.protocol_generator_utils import (
    format_decimal_ru,
    copy_cell_style,
    copy_row_with_styles,
    copy_row_formatting,
    calculate_current_height,
    calculate_header_height,
    copy_column_dimensions,
    A4_HEIGHT_POINTS,
    DEFAULT_ROW_HEIGHT,
)
from ..utils.employee_utils import get_employee_name
from openpyxl.worksheet.header_footer import _HeaderFooterPart

logger = logging.getLogger(__name__)

_CM_TO_INCH = 2.54  # коэффициент для перевода сантиметров в дюймы


def set_sheet_margins(sheet):
    """Устанавливает фиксированные поля страницы на листе Excel."""
    sheet.page_margins.left = 1.5 / _CM_TO_INCH
    sheet.page_margins.right = 1.0 / _CM_TO_INCH
    sheet.page_margins.top = 1.1 / _CM_TO_INCH
    sheet.page_margins.bottom = 0.9 / _CM_TO_INCH


def process_cell_markers(protocol: Protocol, cell_value: str) -> str:
    """
    Обрабатывает все метки в ячейке.
    """
    if not cell_value or not isinstance(cell_value, str):
        return cell_value

    # Быстрая проверка на наличие меток
    if "{" not in cell_value:
        return cell_value

    try:
        result = cell_value

        # Находим все метки сразу
        start = 0
        while True:
            start = result.find("{", start)
            if start == -1:
                break

            end = result.find("}", start)
            if end == -1:
                break

            marker = result[start + 1 : end]

            # Пропускаем метки условий отбора
            if not marker.startswith("sel_cond_") and not marker == "bu":
                value = get_marker_value_title(protocol, marker)
                result = result.replace(f"{{{marker}}}", value)

            start = end + 1

        # Обрабатываем условия отбора один раз в конце
        return process_selection_conditions_row(protocol, result)

    except Exception as e:
        logger.error(f"Ошибка при обработке меток в ячейке: {str(e)}")
        return cell_value


def get_marker_value_title(protocol: Protocol, marker: str) -> str:
    """
    Возвращает значение для метки в заголовке протокола.
    """
    try:
        if marker == "test_protocol_number":
            if not protocol.test_protocol_number:
                return ""

            base_number = protocol.test_protocol_number

            # Если протокол аккредитован, добавляем суффикс
            if protocol.is_accredited:
                # Получаем все объекты испытаний из проб протокола
                test_objects = ", ".join(
                    [
                        sample.test_object
                        for sample in protocol.samples.filter(is_deleted=False)
                    ]
                )

                # Определяем суффикс на основе объектов испытаний
                suffix = ""
                objects = [obj.strip().lower() for obj in test_objects.split(",")]
                for test_object in objects:
                    if "конденсат" in test_object:
                        suffix = "дк"
                        break
                    elif "нефть" in test_object:
                        suffix = "н"
                        break

                base_number = (
                    f"{protocol.test_protocol_number}/07/{suffix}"
                    if suffix
                    else f"{protocol.test_protocol_number}/07"
                )

            # Добавляем дату протокола
            if protocol.test_protocol_date:
                return f"{base_number} от {protocol.test_protocol_date.strftime('%d.%m.%Y')}"
            return base_number

        elif marker == "subd":
            return protocol.branch or ""
        elif marker == "tel":
            return protocol.phone or ""
        elif marker == "res_object":
            # Получаем объекты испытаний из всех проб
            objects = [
                sample.test_object
                for sample in protocol.samples.filter(is_deleted=False)
            ]
            return ", ".join(objects) if objects else ""
        elif marker == "sampling_location":
            locations = [
                sample.sampling_location_detail
                for sample in protocol.samples.filter(is_deleted=False)
                if sample.sampling_location_detail
            ]
            return ", ".join(locations) if locations else ""
        elif marker == "sampling_date":
            dates = sorted(
                set(
                    sample.sampling_date.strftime("%d.%m.%Y")
                    for sample in protocol.samples.filter(is_deleted=False)
                    if sample.sampling_date
                )
            )
            return ", ".join(dates) if dates else ""
        elif marker == "receiving_date":
            dates = sorted(
                set(
                    sample.receiving_date.strftime("%d.%m.%Y")
                    for sample in protocol.samples.filter(is_deleted=False)
                    if sample.receiving_date
                )
            )
            return ", ".join(dates) if dates else ""
        elif marker == "laboratory_activity_dates":
            # Получаем все даты лабораторной деятельности из расчетов
            dates = [
                calc.laboratory_activity_date
                for sample in protocol.samples.filter(is_deleted=False)
                for calc in sample.calculations.filter(is_deleted=False)
                if calc.laboratory_activity_date
            ]
            if dates:
                min_date = min(dates).strftime("%d.%m.%Y")
                max_date = max(dates).strftime("%d.%m.%Y")
                return f"{min_date}-{max_date}" if min_date != max_date else min_date
            return ""
        elif marker == "lab_location":
            # Получаем место осуществления лабораторной деятельности из подразделения или лаборатории
            if protocol.department and protocol.department.laboratory_location:
                return protocol.department.laboratory_location
            elif protocol.laboratory.laboratory_location:
                return protocol.laboratory.laboratory_location
            return ""
        elif marker == "sampling_act_number":
            return protocol.sampling_act_number or ""
        elif marker == "registration_number":
            numbers = [
                sample.registration_number
                for sample in protocol.samples.filter(is_deleted=False)
            ]
            return ", ".join(numbers) if numbers else ""

        return ""
    except Exception as e:
        logger.error(f"Ошибка при получении значения для метки {marker}: {str(e)}")
        return ""


def copy_table_header(
    source_sheet, target_sheet, start_marker, end_marker
) -> tuple[int, int]:
    """
    Копирует шапку таблицы между маркерами в новый лист.
    Возвращает номер строки начала и конца шапки.
    """
    header_start = None
    header_end = None
    current_row_new = 1

    for current_row in range(1, source_sheet.max_row + 1):
        row = list(
            source_sheet.iter_rows(
                min_row=current_row, max_row=current_row, values_only=True
            )
        )[0]

        if header_start is None:
            if any(cell and str(cell).strip() == start_marker for cell in row):
                header_start = current_row
                continue
        else:
            if any(cell and str(cell).strip() == end_marker for cell in row):
                header_end = current_row
                break

            copy_row_with_styles(
                source_sheet, target_sheet, current_row, current_row_new
            )
            current_row_new += 1

    return header_start, header_end


def add_standalone_method(
    calc,
    current_row,
    current_sheet,
    template_sheet,
    table_header_start,
    table_header_end,
    template_row_num,
    merged_cells_map,
    idx,
    sheet_number,
):
    """
    Добавляет одиночный метод в таблицу.
    """
    # Рассчитываем высоту шапки таблицы
    header_height = calculate_header_height(
        template_sheet, table_header_start, table_header_end
    )

    # Проверяем высоту листа: текущая высота + шапка + одна строка данных
    current_height = calculate_current_height(current_sheet)
    if current_height + header_height + DEFAULT_ROW_HEIGHT > A4_HEIGHT_POINTS:
        sheet_number += 1
        current_sheet = current_sheet.parent.create_sheet(f"Лист{sheet_number}")
        set_sheet_margins(current_sheet)
        enforce_fit_to_page(current_sheet)
        current_row = 1

        # Копируем размеры столбцов
        copy_column_dimensions(template_sheet, current_sheet)

        # Создаем новую карту объединенных ячеек для нового листа
        sheet_merged_cells_map = current_sheet.merged_cells

        # Копируем шапку таблицы с сохранением форматирования и объединенных ячеек
        for row_num in range(table_header_start + 1, table_header_end):
            copy_row_formatting(
                template_sheet,
                current_sheet,
                row_num,
                current_row,
                sheet_merged_cells_map,
            )
            current_row += 1
    else:
        # Если это первый элемент и на листе есть место для шапки + данных, копируем шапку
        if (
            idx == 1
            and current_height + header_height + DEFAULT_ROW_HEIGHT <= A4_HEIGHT_POINTS
        ):
            # Копируем шапку таблицы с сохранением форматирования и объединенных ячеек
            for row_num in range(table_header_start + 1, table_header_end):
                copy_row_formatting(
                    template_sheet,
                    current_sheet,
                    row_num,
                    current_row,
                    merged_cells_map,
                )
                current_row += 1
        sheet_merged_cells_map = current_sheet.merged_cells

    # Копируем шаблонную строку с сохранением форматирования и объединенных ячеек
    copy_row_formatting(
        template_sheet,
        current_sheet,
        template_row_num,
        current_row,
        sheet_merged_cells_map,
    )

    # Заполняем значения
    for col in range(1, template_sheet.max_column + 1):
        cell = current_sheet.cell(row=current_row, column=col)
        if not cell.value:
            continue

        value = str(cell.value)
        if "{id_method}" in value:
            cell.value = value.replace("{id_method}", str(idx))
        elif "{name_method}" in value:
            cell.value = value.replace("{name_method}", calc.research_method.name)
        elif "{unit}" in value:
            cell.value = value.replace("{unit}", calc.unit or "не указано")
        elif "{result}" in value:
            cell.value = value.replace("{result}", format_decimal_ru(calc.result))
        elif "{measurement_error}" in value:
            error_value = calc.measurement_error
            formatted_error = (
                error_value
                if error_value and error_value.startswith("-")
                else (
                    f"±{error_value}"
                    if error_value and error_value != "не указано"
                    else "не указано"
                )
            )
            cell.value = value.replace("{measurement_error}", formatted_error)
        elif "{measurement_method}" in value:
            cell.value = value.replace(
                "{measurement_method}",
                calc.research_method.measurement_method or "не указано",
            )

    return current_row + 1, current_sheet, sheet_number


def add_group_methods(
    group_data,
    current_row,
    current_sheet,
    template_sheet,
    table_header_start,
    table_header_end,
    template_row_num,
    merged_cells_map,
    idx,
    sheet_number,
):
    """
    Добавляет группу методов в таблицу.
    """
    # Рассчитываем высоту шапки таблицы
    header_height = calculate_header_height(
        template_sheet, table_header_start, table_header_end
    )

    # Проверяем высоту листа: текущая высота + шапка + все строки группы (заголовок + методы)
    current_height = calculate_current_height(current_sheet)
    group_rows_height = DEFAULT_ROW_HEIGHT * (
        len(group_data["calculations"]) + 1
    )  # +1 для заголовка группы
    if current_height + header_height + group_rows_height > A4_HEIGHT_POINTS:
        sheet_number += 1
        current_sheet = current_sheet.parent.create_sheet(f"Лист{sheet_number}")
        set_sheet_margins(current_sheet)
        enforce_fit_to_page(current_sheet)
        current_row = 1

        # Копируем размеры столбцов
        copy_column_dimensions(template_sheet, current_sheet)

        # Создаем новую карту объединенных ячеек для нового листа
        sheet_merged_cells_map = current_sheet.merged_cells

        # Копируем шапку таблицы с сохранением форматирования и объединенных ячеек
        for row_num in range(table_header_start + 1, table_header_end):
            copy_row_formatting(
                template_sheet,
                current_sheet,
                row_num,
                current_row,
                sheet_merged_cells_map,
            )
            current_row += 1
    else:
        # Если это первый элемент и на листе есть место для шапки + данных, копируем шапку
        if (
            idx == 1
            and current_height + header_height + group_rows_height <= A4_HEIGHT_POINTS
        ):
            # Копируем шапку таблицы с сохранением форматирования и объединенных ячеек
            for row_num in range(table_header_start + 1, table_header_end):
                copy_row_formatting(
                    template_sheet,
                    current_sheet,
                    row_num,
                    current_row,
                    merged_cells_map,
                )
                current_row += 1
        sheet_merged_cells_map = current_sheet.merged_cells

    # Получаем общие данные для группы
    measurement_methods = set(
        method.measurement_method for method in group_data["methods"]
    )
    units = set(calc.unit for calc in group_data["calculations"])

    common_measurement_method = (
        next(iter(measurement_methods)) if len(measurement_methods) == 1 else None
    )
    common_unit = next(iter(units)) if len(units) == 1 else None

    # Копируем шаблонную строку для заголовка группы с сохранением форматирования и объединенных ячеек
    copy_row_formatting(
        template_sheet,
        current_sheet,
        template_row_num,
        current_row,
        sheet_merged_cells_map,
    )

    # Заполняем заголовок группы
    for col in range(1, template_sheet.max_column + 1):
        cell = current_sheet.cell(row=current_row, column=col)

        # Убираем нижнюю границу
        if cell.border:
            new_border = copy(cell.border)
            new_border.bottom = None
            cell.border = new_border

        if not cell.value:
            continue

        value = str(cell.value)
        if "{id_method}" in value:
            cell.value = value.replace("{id_method}", str(idx))
        elif "{name_method}" in value:
            cell.value = value.replace("{name_method}", group_data["name"])
        elif "{unit}" in value:
            cell.value = value.replace("{unit}", common_unit or "не указано")
        elif "{measurement_method}" in value:
            cell.value = value.replace(
                "{measurement_method}",
                common_measurement_method or "не указано",
            )
        else:
            # Для остальных полей в заголовке группы ставим пустые значения
            for placeholder in ["{result}", "{measurement_error}"]:
                if placeholder in value:
                    cell.value = value.replace(placeholder, "")

    current_row += 1

    # Добавляем методы группы
    for i, calc in enumerate(group_data["calculations"]):
        # Копируем шаблонную строку для метода с сохранением форматирования и объединенных ячеек
        copy_row_formatting(
            template_sheet,
            current_sheet,
            template_row_num,
            current_row,
            sheet_merged_cells_map,
        )

        # Особая обработка границ
        last_method_in_group = i == len(group_data["calculations"]) - 1
        for col in range(1, template_sheet.max_column + 1):
            cell = current_sheet.cell(row=current_row, column=col)
            if cell.border:
                new_border = copy(cell.border)
                new_border.top = None  # Всегда убираем верхнюю границу

                if not last_method_in_group:
                    new_border.bottom = (
                        None  # Убираем нижнюю границу для всех кроме последнего
                    )
                cell.border = new_border

        # Заполняем значения метода
        for col in range(1, template_sheet.max_column + 1):
            cell = current_sheet.cell(row=current_row, column=col)
            if not cell.value:
                continue

            value = str(cell.value)
            if "{id_method}" in value:
                cell.value = ""  # Пустой ID для методов в группе
            elif "{name_method}" in value:
                # Делаем первую букву метода строчной
                method_name = calc.research_method.name
                if method_name:
                    method_name = method_name[0].lower() + method_name[1:]
                    cell.value = method_name
            elif "{result}" in value:
                cell.value = value.replace("{result}", format_decimal_ru(calc.result))
            elif "{measurement_error}" in value:
                error_value = calc.measurement_error
                formatted_error = (
                    error_value
                    if error_value and error_value.startswith("-")
                    else (
                        f"±{error_value}"
                        if error_value and error_value != "не указано"
                        else "не указано"
                    )
                )
                cell.value = value.replace("{measurement_error}", formatted_error)
            elif "{unit}" in value:
                cell.value = ""  # Пустая единица измерения для методов в группе
            elif "{measurement_method}" in value:
                cell.value = ""  # Пустой метод измерения для методов в группе

        current_row += 1

    return current_row, current_sheet, sheet_number


def process_selection_conditions_row(protocol, cell_value: str) -> str:
    """
    Обрабатывает метки условий отбора в ячейке.
    Если условий отбора нет в шаблоне, ставит "б/у" и пропускает строки с метками условий.
    """
    if not cell_value or not isinstance(cell_value, str):
        return cell_value

    # Быстрая проверка на наличие меток
    if "{sel_cond_" not in cell_value and "{bu}" not in cell_value:
        return cell_value

    # Проверяем наличие условий отбора
    selection_conditions = protocol.selection_conditions or {}

    # Проверяем, есть ли хотя бы одно не-null значение
    has_conditions = any(
        value is not None and value != "" for value in selection_conditions.values()
    )

    # Если есть метка {bu} и нет значений условий отбора, заменяем на "б/у"
    if "{bu}" in cell_value and not has_conditions:
        return cell_value.replace("{bu}", "б/у")

    # Если нет значений условий отбора и это строка с метками условий, пропускаем её
    if not has_conditions and any(
        f"{{sel_cond_{type}_{i}}}" in cell_value
        for type in ["name", "val", "unit"]
        for i in range(1, 6)
    ):
        return None

    # Если есть условия отбора, обрабатываем их
    if has_conditions:
        # Получаем единицы измерения из шаблона протокола
        template_conditions = (
            protocol.excel_template.selection_conditions
            if protocol.excel_template
            else []
        )
        units_map = (
            {item["name"]: item["unit"] for item in template_conditions}
            if template_conditions
            else {}
        )

        # Преобразуем selection_conditions в список, только не-null значения
        selection_conditions_list = []
        for name, value in selection_conditions.items():
            if value is not None and value != "":
                selection_conditions_list.append(
                    {"name": name, "value": value, "unit": units_map.get(name, "")}
                )

        result = cell_value

        # Проверяем наличие любых меток условий в строке
        has_condition_tags = any(
            f"{{sel_cond_{type}_{i}}}" in result
            for type in ["name", "val", "unit"]
            for i in range(1, 6)
        )

        # Если в строке есть метки условий, но нет значений для заполнения - пропускаем строку
        if has_condition_tags and not selection_conditions_list:
            return None

        # Определяем индекс для текущей строки
        current_index = None
        for i in range(1, 6):
            name_tag = f"{{sel_cond_name_{i}}}"
            val_tag = f"{{sel_cond_val_{i}}}"
            unit_tag = f"{{sel_cond_unit_{i}}}"

            # Если в текущей строке есть хотя бы одна метка с этим индексом
            if name_tag in result or val_tag in result or unit_tag in result:
                current_index = i
                break

        # Если нашли индекс и он больше количества значений - пропускаем строку
        if current_index is not None:
            if current_index > len(selection_conditions_list):
                return None

            # Заполняем значения для текущего индекса
            condition = selection_conditions_list[current_index - 1]
            name_tag = f"{{sel_cond_name_{current_index}}}"
            val_tag = f"{{sel_cond_val_{current_index}}}"
            unit_tag = f"{{sel_cond_unit_{current_index}}}"

            if name_tag in result:
                result = result.replace(name_tag, condition["name"] + ":")
            if val_tag in result:
                formatted_value = format_decimal_ru(condition["value"])
                result = result.replace(val_tag, formatted_value)
            if unit_tag in result:
                result = result.replace(unit_tag, condition["unit"])

        # Обрабатываем метку {bu} - если есть условия, просто убираем метку
        if "{bu}" in result:
            result = result.replace("{bu}", "")

        return result

    return cell_value


def check_method_name(method_name, test_objects):
    """
    Проверяет, соответствует ли метод объекту испытаний
    """
    method_lower = method_name.lower()
    test_objects_lower = [obj.lower() for obj in test_objects]

    logger.debug(f"Проверка метода: {method_name}")
    logger.debug(f"Объекты испытаний: {', '.join(test_objects)}")

    # Если в методе есть "нефть", но в объектах испытаний нет нефти
    if "нефть" in method_lower and not any(
        "нефть" in obj for obj in test_objects_lower
    ):
        logger.info(
            f"Метод '{method_name}' содержит слово 'нефть', но в объектах испытаний нет нефти"
        )
        return False

    # Если в методе есть "конденсат", проверяем наличие конденсата в объектах испытаний
    if "конденсат" in method_lower:
        has_condensate = any("конденсат" in obj for obj in test_objects_lower)
        if has_condensate:
            logger.debug(
                f"Метод '{method_name}' подходит для объектов испытаний (найден конденсат)"
            )
            return True
        else:
            logger.info(
                f"Метод '{method_name}' содержит слово 'конденсат', но в объектах испытаний нет конденсата"
            )
            return False

    # Если метод не содержит специфических слов "нефть" или "конденсат" - считаем что он подходит
    logger.debug(
        f"Метод '{method_name}' прошел проверку на соответствие объектам испытаний (общий метод)"
    )
    return True


def process_header(protocol, template_sheet, new_sheet, merged_cells_map):
    """
    Обрабатывает шапку протокола.
    """
    current_row_new = 1
    found_start = False

    # Идем по строкам шаблона, ищем начало шапки
    for current_row in range(1, template_sheet.max_row + 1):
        row = list(
            template_sheet.iter_rows(
                min_row=current_row, max_row=current_row, values_only=True
            )
        )[0]

        # Ждем начало шапки
        if not found_start:
            if any(cell and str(cell).strip() == "{{start_header}}" for cell in row):
                found_start = True
            continue

        # Если дошли до конца шапки - выходим
        if any(cell and str(cell).strip() == "{{end_header}}" for cell in row):
            return current_row + 1  # Возвращаем следующую строку после метки

        # Пропускаем строку аккредитации если нужно
        if (
            not protocol.is_accredited
            and protocol.excel_template
            and current_row == protocol.excel_template.accreditation_header_row
        ):
            continue

        # Копируем размеры строк
        if current_row in template_sheet.row_dimensions:
            new_sheet.row_dimensions[current_row_new] = copy(
                template_sheet.row_dimensions[current_row]
            )

        # Копируем объединенные ячейки для текущей строки
        for merged_range in template_sheet.merged_cells.ranges:
            if merged_range.min_row == current_row:
                # Создаем новый диапазон с учетом смещения строк
                new_range = openpyxl.worksheet.cell_range.CellRange(
                    min_col=merged_range.min_col,
                    min_row=current_row_new,
                    max_col=merged_range.max_col,
                    max_row=current_row_new
                    + (merged_range.max_row - merged_range.min_row),
                )
                merged_cells_map.add(new_range)

        copy_row_with_styles(template_sheet, new_sheet, current_row, current_row_new)

        # Обрабатываем метки в скопированной строке
        skip_row = False
        for col in range(1, template_sheet.max_column + 1):
            cell = new_sheet.cell(row=current_row_new, column=col)
            if cell.value:
                processed_value = process_cell_markers(protocol, str(cell.value))
                if processed_value is None:
                    skip_row = True
                    break
                cell.value = processed_value

        if skip_row:
            # Удаляем размеры строки и объединенные ячейки
            if current_row_new in new_sheet.row_dimensions:
                del new_sheet.row_dimensions[current_row_new]
            for merged_range in list(new_sheet.merged_cells.ranges):
                if merged_range.min_row == current_row_new:
                    new_sheet.merged_cells.remove(merged_range)
            continue

        current_row_new += 1

    return current_row


def process_header_and_conditions(
    protocol, template_sheet, new_sheet, start_row, merged_cells_map
):
    """
    Обрабатывает заголовок и условия отбора после шапки до начала таблицы.
    """
    current_row_new = new_sheet.max_row + 1

    # Продолжаем со строки после шапки
    for current_row in range(start_row, template_sheet.max_row + 1):
        row = list(
            template_sheet.iter_rows(
                min_row=current_row, max_row=current_row, values_only=True
            )
        )[0]

        # Если дошли до начала таблицы - выходим
        if any(cell and str(cell).strip() == "{{start_table1}}" for cell in row):
            return current_row

        # Проверяем, нужно ли пропустить строку
        skip_row = False
        processed_values = []

        # Обрабатываем метки в строке перед копированием
        for cell_value in row:
            if cell_value:
                processed_value = process_cell_markers(protocol, str(cell_value))
                if processed_value is None:
                    skip_row = True
                    break
                processed_values.append(processed_value)
            else:
                processed_values.append(cell_value)

        if skip_row:
            continue

        if current_row in template_sheet.row_dimensions:
            new_sheet.row_dimensions[current_row_new] = copy(
                template_sheet.row_dimensions[current_row]
            )

        for merged_range in template_sheet.merged_cells.ranges:
            if merged_range.min_row == current_row:
                # Создаем новый диапазон с учетом смещения строк
                new_range = openpyxl.worksheet.cell_range.CellRange(
                    min_col=merged_range.min_col,
                    min_row=current_row_new,
                    max_col=merged_range.max_col,
                    max_row=current_row_new
                    + (merged_range.max_row - merged_range.min_row),
                )
                merged_cells_map.add(new_range)

        copy_row_with_styles(template_sheet, new_sheet, current_row, current_row_new)

        # Заполняем обработанные значения
        for col, processed_value in enumerate(processed_values, start=1):
            if (
                processed_value is not None
            ):  # Проверяем на None, чтобы не заполнены стили пустым значением
                cell = new_sheet.cell(row=current_row_new, column=col)
                cell.value = processed_value

        current_row_new += 1

    return current_row


def process_methods_table(
    protocol,
    template_sheet,
    new_sheet,
    table_start,
    merged_cells_map,
    current_row,
    sheet_number,
):
    """
    Обрабатывает таблицу с методами исследования
    """
    current_sheet = new_sheet
    workbook = new_sheet.parent

    # Получаем объекты испытаний из проб протокола
    test_objects = []
    for sample in protocol.samples.filter(is_deleted=False):
        if sample.test_object:
            test_objects.append(sample.test_object)

    logger.info(f"Объекты испытаний: {', '.join(test_objects)}")

    # Получаем все расчеты для протокола
    calculations = []
    for sample in protocol.samples.filter(is_deleted=False):
        calcs = (
            sample.calculations.filter(is_deleted=False)
            .select_related("research_method")
            .annotate(
                method_sort_order=models.Subquery(
                    ResearchObjectMethod.objects.filter(
                        research_method=models.OuterRef("research_method"),
                        research_object__laboratory=protocol.laboratory,
                        research_object__department=protocol.department,
                        is_deleted=False,
                    ).values("sort_order")[:1]
                )
            )
            .order_by("method_sort_order", "research_method__name", "created_at")
        )
        calculations.extend(calcs)

    # Фильтруем методы по соответствию объектам испытаний
    valid_calculations = []
    for calc in calculations:
        if check_method_name(calc.research_method.name, test_objects):
            valid_calculations.append(calc)

    # Если нет методов для вывода - пропускаем таблицу
    if not valid_calculations:
        logger.info("Нет методов для вывода в таблице 1, пропускаем таблицу")
        return current_sheet

    # Ищем начало и конец таблицы
    table_header_start = None
    table_header_end = None
    table_end = None

    for row_num in range(table_start, template_sheet.max_row + 1):
        row = list(
            template_sheet.iter_rows(min_row=row_num, max_row=row_num, values_only=True)
        )[0]

        # Проверяем метки таблицы
        for cell in row:
            if cell:
                cell_str = str(cell).strip()
                if table_header_start is None and cell_str == "{{start_table1}}":
                    table_header_start = row_num
                    break
                elif table_header_start is not None and cell_str == "{start_table1}":
                    table_header_end = row_num
                    break
                elif cell_str == "{{end_table1}}":
                    table_end = row_num
                    break

    # Если не нашли начало таблицы - пропускаем
    if table_header_start is None:
        logger.info("Не найдена метка {{start_table1}} в шаблоне, пропускаем таблицу")
        return current_sheet

    # Если не нашли конец шапки - используем следующую строку после начала
    if table_header_end is None:
        table_header_end = table_header_start + 1
        logger.info(
            "Не найдена метка {start_table1}, используем следующую строку после {{start_table1}}"
        )

    # Находим шаблонную строку между {start_table1} и {end_table1}
    template_row = None
    template_row_num = None

    # Ищем шаблонную строку начиная с позиции table_header_end
    for row_num in range(table_header_end, template_sheet.max_row + 1):
        row = list(
            template_sheet.iter_rows(min_row=row_num, max_row=row_num, values_only=True)
        )[0]

        if any(cell and str(cell).strip() == "{start_table1}" for cell in row):
            template_row_num = row_num + 1
            template_row = list(
                template_sheet.iter_rows(
                    min_row=template_row_num, max_row=template_row_num, values_only=True
                )
            )[0]
            break
        elif any(cell and str(cell).strip() == "{end_table1}" for cell in row):
            break

    # Если не нашли шаблонную строку - пропускаем таблицу
    if not template_row:
        logger.info("Не найдена шаблонная строка таблицы, пропускаем таблицу")
        return current_sheet

    # Группируем методы, но сохраняем порядок
    grouped_calculations = {}
    processed_calculations = []  # Список для сохранения порядка методов

    for calc in valid_calculations:
        # Проверяем, входит ли метод в группу
        if (
            calc.research_method.is_group_member
            and calc.research_method.groups.exists()
        ):
            group = calc.research_method.groups.first()
            group_id = group.id
            group_name = group.name

            if group_id not in grouped_calculations:
                grouped_calculations[group_id] = {
                    "name": group_name,
                    "calculations": [],
                    "methods": [],
                }

            grouped_calculations[group_id]["calculations"].append(calc)
            grouped_calculations[group_id]["methods"].append(calc.research_method)
            processed_calculations.append(
                {"type": "group", "group_id": group_id, "calc": calc}
            )
        else:
            processed_calculations.append({"type": "standalone", "calc": calc})

    # Счетчик для нумерации методов
    idx = 1

    # Обрабатываем методы в порядке их следования
    current_group = None
    group_methods = []

    for item in processed_calculations:
        if item["type"] == "standalone":
            # Если перед этим была группа, выводим её
            if group_methods:
                group_id = current_group
                group_data = grouped_calculations[group_id]

                # Проверяем, содержит ли группа методы с "нефть" или "конденсат"
                has_special_methods = any(
                    "нефть" in method.name.lower() or "конденсат" in method.name.lower()
                    for method in group_data["methods"]
                )

                if has_special_methods:
                    # Создаем копию первого расчета группы
                    group_calc = copy(group_data["calculations"][0])
                    # Меняем название метода на название группы
                    group_calc.research_method.name = group_data["name"]
                    current_row, current_sheet, sheet_number = add_standalone_method(
                        group_calc,
                        current_row,
                        current_sheet,
                        template_sheet,
                        table_header_start,
                        table_header_end,
                        template_row_num,
                        merged_cells_map,
                        idx,
                        sheet_number,
                    )
                else:
                    # Добавляем группу как есть
                    current_row, current_sheet, sheet_number = add_group_methods(
                        group_data,
                        current_row,
                        current_sheet,
                        template_sheet,
                        table_header_start,
                        table_header_end,
                        template_row_num,
                        merged_cells_map,
                        idx,
                        sheet_number,
                    )
                idx += 1
                group_methods = []
                current_group = None

            # Добавляем одиночный метод
            current_row, current_sheet, sheet_number = add_standalone_method(
                item["calc"],
                current_row,
                current_sheet,
                template_sheet,
                table_header_start,
                table_header_end,
                template_row_num,
                merged_cells_map,
                idx,
                sheet_number,
            )
            idx += 1
        else:  # type == "group"
            if current_group is None:
                current_group = item["group_id"]
                group_methods = [item]
            elif current_group == item["group_id"]:
                group_methods.append(item)
            else:
                # Выводим предыдущую группу
                group_data = grouped_calculations[current_group]

                # Проверяем, содержит ли группа методы с "нефть" или "конденсат"
                has_special_methods = any(
                    "нефть" in method.name.lower() or "конденсат" in method.name.lower()
                    for method in group_data["methods"]
                )

                if has_special_methods:
                    # Создаем копию первого расчета группы
                    group_calc = copy(group_data["calculations"][0])
                    # Меняем название метода на название группы
                    group_calc.research_method.name = group_data["name"]
                    current_row, current_sheet, sheet_number = add_standalone_method(
                        group_calc,
                        current_row,
                        current_sheet,
                        template_sheet,
                        table_header_start,
                        table_header_end,
                        template_row_num,
                        merged_cells_map,
                        idx,
                        sheet_number,
                    )
                else:
                    # Добавляем группу как есть
                    current_row, current_sheet, sheet_number = add_group_methods(
                        group_data,
                        current_row,
                        current_sheet,
                        template_sheet,
                        table_header_start,
                        table_header_end,
                        template_row_num,
                        merged_cells_map,
                        idx,
                        sheet_number,
                    )
                idx += 1

                # Начинаем новую группу
                current_group = item["group_id"]
                group_methods = [item]

    # Обрабатываем последнюю группу, если она есть
    if group_methods:
        group_data = grouped_calculations[current_group]

        # Проверяем, содержит ли группа методы с "нефть" или "конденсат"
        has_special_methods = any(
            "нефть" in method.name.lower() or "конденсат" in method.name.lower()
            for method in group_data["methods"]
        )

        if has_special_methods:
            # Создаем копию первого расчета группы
            group_calc = copy(group_data["calculations"][0])
            # Меняем название метода на название группы
            group_calc.research_method.name = group_data["name"]
            current_row, current_sheet, sheet_number = add_standalone_method(
                group_calc,
                current_row,
                current_sheet,
                template_sheet,
                table_header_start,
                table_header_end,
                template_row_num,
                merged_cells_map,
                idx,
                sheet_number,
            )
        else:
            # Добавляем группу как есть
            current_row, current_sheet, sheet_number = add_group_methods(
                group_data,
                current_row,
                current_sheet,
                template_sheet,
                table_header_start,
                table_header_end,
                template_row_num,
                merged_cells_map,
                idx,
                sheet_number,
            )

    return current_sheet


def process_equipment_table(
    protocol,
    template_sheet,
    current_sheet,
    table_start,
    merged_cells_map,
    current_row,
    sheet_number,
):
    """
    Обрабатывает таблицу с оборудованием
    """
    logger.info(f"Начинаем поиск меток таблицы оборудования с строки {table_start}")

    # Получаем уникальное оборудование из всех расчетов протокола
    equipment_ids = set()
    for sample in protocol.samples.filter(is_deleted=False):
        for calc in sample.calculations.filter(is_deleted=False):
            if calc.equipment_data:
                # Обрабатываем список словарей с ID оборудования
                if isinstance(calc.equipment_data, list):
                    for item in calc.equipment_data:
                        if isinstance(item, dict) and "id" in item:
                            equipment_ids.add(item["id"])
                        elif isinstance(item, (int, str)):
                            equipment_ids.add(int(item))
                elif (
                    isinstance(calc.equipment_data, dict)
                    and "id" in calc.equipment_data
                ):
                    equipment_ids.add(calc.equipment_data["id"])
                elif isinstance(calc.equipment_data, (int, str)):
                    equipment_ids.add(int(calc.equipment_data))

    # Получаем оборудование
    equipment_list = Equipment.objects.filter(
        id__in=list(equipment_ids), is_deleted=False
    ).order_by("name", "version")

    # Если нет оборудования - пропускаем таблицу
    if not equipment_list.exists():
        logger.info("Нет оборудования для вывода в таблице 2, пропускаем таблицу")
        return current_sheet

    # Находим шапку таблицы оборудования
    table_header_start = None
    table_header_end = None
    table_end = None

    for row_num in range(table_start, template_sheet.max_row + 1):
        row = list(
            template_sheet.iter_rows(min_row=row_num, max_row=row_num, values_only=True)
        )[0]

        cell_values = [str(cell).strip() if cell else "" for cell in row]
        logger.debug(f"Строка {row_num}: {cell_values}")

        if table_header_start is None:
            if any(cell and str(cell).strip() == "{{start_table2}}" for cell in row):
                table_header_start = row_num
                logger.info(
                    f"Найдена метка начала таблицы оборудования в строке {row_num}"
                )
                continue
        else:
            if any(cell and str(cell).strip() == "{start_table2}" for cell in row):
                table_header_end = row_num
                logger.info(
                    f"Найдена метка конца шапки таблицы оборудования в строке {row_num}"
                )
                break
            elif any(cell and str(cell).strip() == "{{end_table2}}" for cell in row):
                table_end = row_num
                break

    # Если не нашли начало таблицы - пропускаем
    if table_header_start is None:
        logger.info("Не найдена метка {{start_table2}} в шаблоне, пропускаем таблицу")
        return current_sheet

    # Если не нашли конец шапки - используем следующую строку после начала
    if table_header_end is None:
        table_header_end = table_header_start + 1
        logger.info(
            "Не найдена метка {start_table2}, используем следующую строку после {{start_table2}}"
        )

    # Находим шаблонную строку
    template_row = None
    template_row_num = None

    for row_num in range(table_header_end, template_sheet.max_row + 1):
        row = list(
            template_sheet.iter_rows(min_row=row_num, max_row=row_num, values_only=True)
        )[0]

        if any(cell and str(cell).strip() == "{start_table2}" for cell in row):
            template_row_num = row_num + 1
            template_row = list(
                template_sheet.iter_rows(
                    min_row=template_row_num, max_row=template_row_num, values_only=True
                )
            )[0]
            break
        elif any(cell and str(cell).strip() == "{end_table2}" for cell in row):
            break

    # Если не нашли шаблонную строку - пропускаем таблицу
    if not template_row:
        logger.info(
            "Не найдена шаблонная строка таблицы оборудования, пропускаем таблицу"
        )
        return current_sheet

    # Создаем новую карту объединенных ячеек для каждого листа
    sheet_merged_cells_map = current_sheet.merged_cells

    # Добавляем строки с оборудованием
    idx = 1
    for equipment in equipment_list:
        # Рассчитываем высоту шапки таблицы
        header_height = calculate_header_height(
            template_sheet, table_header_start, table_header_end
        )

        # Проверяем высоту листа: текущая высота + шапка + одна строка данных
        current_height = calculate_current_height(current_sheet)
        if current_height + header_height + DEFAULT_ROW_HEIGHT > A4_HEIGHT_POINTS:
            sheet_number += 1
            current_sheet = current_sheet.parent.create_sheet(f"Лист{sheet_number}")
            set_sheet_margins(current_sheet)
            enforce_fit_to_page(current_sheet)
            current_row = 1

            # Копируем размеры столбцов
            copy_column_dimensions(template_sheet, current_sheet)

            # Создаем новую карту объединенных ячеек для нового листа
            sheet_merged_cells_map = current_sheet.merged_cells

            # Копируем шапку таблицы с сохранением форматирования и объединенных ячеек
            for row_num in range(table_header_start + 1, table_header_end):
                copy_row_formatting(
                    template_sheet,
                    current_sheet,
                    row_num,
                    current_row,
                    sheet_merged_cells_map,
                )
                current_row += 1
        else:
            # Если это первый элемент и на листе есть место для шапки + данных, копируем шапку
            if (
                idx == 1
                and current_height + header_height + DEFAULT_ROW_HEIGHT
                <= A4_HEIGHT_POINTS
            ):
                # Копируем шапку таблицы с сохранением форматирования и объединенных ячеек
                for row_num in range(table_header_start + 1, table_header_end):
                    copy_row_formatting(
                        template_sheet,
                        current_sheet,
                        row_num,
                        current_row,
                        sheet_merged_cells_map,
                    )
                    current_row += 1

        # Копируем шаблонную строку с сохранением форматирования и объединенных ячеек
        copy_row_formatting(
            template_sheet,
            current_sheet,
            template_row_num,
            current_row,
            sheet_merged_cells_map,
        )

        # Заполняем значения
        for col in range(1, template_sheet.max_column + 1):
            cell = current_sheet.cell(row=current_row, column=col)
            if not cell.value:
                continue

            value = str(cell.value)
            if "{id_equipment}" in value:
                cell.value = value.replace("{id_equipment}", str(idx))
            elif "{name_equipment}" in value:
                cell.value = value.replace("{name_equipment}", equipment.name)
            elif "{serial_num}" in value:
                cell.value = value.replace("{serial_num}", equipment.serial_number)
            elif "{ver_info}" in value:
                cell.value = value.replace("{ver_info}", equipment.verification_info)
            elif "{ver_date}" in value:
                formatted_date = formats.date_format(
                    equipment.verification_date, "d.m.Y"
                )
                cell.value = value.replace("{ver_date}", formatted_date)
            elif "{ver_end_date}" in value:
                formatted_date = formats.date_format(
                    equipment.verification_end_date, "d.m.Y"
                )
                cell.value = value.replace("{ver_end_date}", formatted_date)

        current_row += 1
        idx += 1

    return current_sheet


def process_nd_table(
    protocol,
    template_sheet,
    current_sheet,
    table_start,
    merged_cells_map,
    current_row,
    sheet_number,
):
    """
    Обрабатывает таблицу с нормативными документами
    """
    # Получаем все расчеты для протокола в нужном порядке
    calculations = []
    for sample in protocol.samples.filter(is_deleted=False):
        calcs = (
            sample.calculations.filter(is_deleted=False)
            .select_related("research_method")
            .annotate(
                method_sort_order=models.Subquery(
                    ResearchObjectMethod.objects.filter(
                        research_method=models.OuterRef("research_method"),
                        research_object__laboratory=protocol.laboratory,
                        research_object__department=protocol.department,
                        is_deleted=False,
                    ).values("sort_order")[:1]
                )
            )
            .order_by("method_sort_order", "research_method__name", "created_at")
        )
        calculations.extend(calcs)

    # Получаем объекты испытаний из проб протокола
    test_objects = []
    for sample in protocol.samples.filter(is_deleted=False):
        if sample.test_object:
            test_objects.append(sample.test_object)

    # Собираем НД в порядке методов
    nd_list = []
    seen_nd = set()  # Для отслеживания уникальных НД

    for calc in calculations:
        # Проверяем соответствие метода объекту испытаний
        if not check_method_name(calc.research_method.name, test_objects):
            continue

        nd_key = (calc.research_method.nd_code, calc.research_method.nd_name)
        if nd_key not in seen_nd:
            seen_nd.add(nd_key)
            nd_list.append(nd_key)

    # Если нет НД - пропускаем таблицу
    if not nd_list:
        logger.info(
            "Нет нормативных документов для вывода в таблице 3, пропускаем таблицу"
        )
        return current_sheet

    # Находим шапку таблицы НД
    table_header_start = None
    table_header_end = None
    table_end = None

    for row_num in range(table_start, template_sheet.max_row + 1):
        row = list(
            template_sheet.iter_rows(min_row=row_num, max_row=row_num, values_only=True)
        )[0]

        if table_header_start is None:
            if any(cell and str(cell).strip() == "{{start_table3}}" for cell in row):
                table_header_start = row_num
                continue
        else:
            if any(cell and str(cell).strip() == "{start_table3}" for cell in row):
                table_header_end = row_num
                break
            elif any(cell and str(cell).strip() == "{{end_table3}}" for cell in row):
                table_end = row_num
                break

    # Если не нашли начало таблицы - пропускаем
    if table_header_start is None:
        logger.info("Не найдена метка {{start_table3}} в шаблоне, пропускаем таблицу")
        return current_sheet

    # Если не нашли конец шапки - используем следующую строку после начала
    if table_header_end is None:
        table_header_end = table_header_start + 1
        logger.info(
            "Не найдена метка {start_table3}, используем следующую строку после {{start_table3}}"
        )

    # Создаем новую карту объединенных ячеек для каждого листа
    sheet_merged_cells_map = current_sheet.merged_cells

    # Шапка таблицы будет скопирована только при создании нового листа

    # Находим шаблонную строку
    template_row = None
    template_row_num = None

    for row_num in range(table_header_end, template_sheet.max_row + 1):
        row = list(
            template_sheet.iter_rows(min_row=row_num, max_row=row_num, values_only=True)
        )[0]

        if any(cell and str(cell).strip() == "{start_table3}" for cell in row):
            template_row_num = row_num + 1
            template_row = list(
                template_sheet.iter_rows(
                    min_row=template_row_num, max_row=template_row_num, values_only=True
                )
            )[0]
            break
        elif any(cell and str(cell).strip() == "{end_table3}" for cell in row):
            break

    # Если не нашли шаблонную строку - пропускаем таблицу
    if not template_row:
        logger.info("Не найдена шаблонная строка таблицы НД, пропускаем таблицу")
        return current_sheet

    # Добавляем строки с НД
    idx = 1
    for nd_code, nd_name in nd_list:
        # Рассчитываем высоту шапки таблицы
        header_height = calculate_header_height(
            template_sheet, table_header_start, table_header_end
        )

        # Проверяем высоту листа: текущая высота + шапка + одна строка данных
        current_height = calculate_current_height(current_sheet)
        if current_height + header_height + DEFAULT_ROW_HEIGHT > A4_HEIGHT_POINTS:
            sheet_number += 1
            current_sheet = current_sheet.parent.create_sheet(f"Лист{sheet_number}")
            set_sheet_margins(current_sheet)
            enforce_fit_to_page(current_sheet)
            current_row = 1

            # Копируем размеры столбцов
            copy_column_dimensions(template_sheet, current_sheet)

            # Создаем новую карту объединенных ячеек для нового листа
            sheet_merged_cells_map = current_sheet.merged_cells

            # Копируем шапку таблицы с сохранением форматирования и объединенных ячеек
            for row_num in range(table_header_start + 1, table_header_end):
                copy_row_formatting(
                    template_sheet,
                    current_sheet,
                    row_num,
                    current_row,
                    sheet_merged_cells_map,
                )
                current_row += 1
        else:
            # Если это первый элемент и на листе есть место для шапки + данных, копируем шапку
            if (
                idx == 1
                and current_height + header_height + DEFAULT_ROW_HEIGHT
                <= A4_HEIGHT_POINTS
            ):
                # Копируем шапку таблицы с сохранением форматирования и объединенных ячеек
                for row_num in range(table_header_start + 1, table_header_end):
                    copy_row_formatting(
                        template_sheet,
                        current_sheet,
                        row_num,
                        current_row,
                        sheet_merged_cells_map,
                    )
                    current_row += 1

        # Копируем шаблонную строку с сохранением форматирования и объединенных ячеек
        copy_row_formatting(
            template_sheet,
            current_sheet,
            template_row_num,
            current_row,
            sheet_merged_cells_map,
        )

        # Заполняем значения
        for col in range(1, template_sheet.max_column + 1):
            cell = current_sheet.cell(row=current_row, column=col)
            if not cell.value:
                continue

            value = str(cell.value)
            if "{id_nd}" in value:
                cell.value = value.replace("{id_nd}", str(idx))
            elif "{nd_code}" in value:
                cell.value = value.replace("{nd_code}", nd_code)
            elif "{name_nd}" in value:
                cell.value = value.replace("{name_nd}", nd_name)

        current_row += 1
        idx += 1

    return current_sheet


def process_between_tables(
    protocol,
    template_sheet,
    current_sheet,
    table_end,
    merged_cells_map,
    current_row,
    sheet_number,
):
    """
    Обрабатывает данные между таблицами.
    """
    # Создаем новую карту объединенных ячеек для каждого листа
    sheet_merged_cells_map = current_sheet.merged_cells

    # Ищем следующую таблицу
    next_table_start = None
    for row_num in range(table_end + 1, template_sheet.max_row + 1):
        row = list(
            template_sheet.iter_rows(min_row=row_num, max_row=row_num, values_only=True)
        )[0]
        if any(cell and str(cell).strip().startswith("{{start_table") for cell in row):
            next_table_start = row_num
            break

    if not next_table_start:
        return current_sheet, current_row

    # Получаем уникальные ФИО исполнителей из hashMd5
    executors_cache = set()
    for sample in protocol.samples.filter(is_deleted=False):
        for calc in sample.calculations.filter(is_deleted=False):
            if calc.executor:
                name = get_employee_name(calc.executor)
                if name:
                    executors_cache.add(name)

    executors = sorted(executors_cache) if executors_cache else []

    # Копируем и обрабатываем строки между таблицами
    row_with_executor = None
    executor_column = None

    # Сначала найдем строку с меткой {executor}
    for row_num in range(table_end + 1, next_table_start):
        row = list(
            template_sheet.iter_rows(min_row=row_num, max_row=row_num, values_only=True)
        )[0]

        # Пропускаем строки с метками конца таблиц
        if any(
            cell
            and str(cell).strip()
            in [
                "{end_table1}",
                "{end_table2}",
                "{end_table3}",
                "{{end_table1}}",
                "{{end_table2}}",
                "{{end_table3}}",
            ]
            for cell in row
        ):
            continue

        # Проверяем, есть ли в строке метка {executor}
        for col_idx, cell_value in enumerate(row, 1):
            if cell_value and "{executor}" in str(cell_value):
                row_with_executor = row_num
                executor_column = col_idx
                break

        if row_with_executor:
            break

    # Обрабатываем все строки
    for row_num in range(table_end + 1, next_table_start):
        row = list(
            template_sheet.iter_rows(min_row=row_num, max_row=row_num, values_only=True)
        )[0]

        # Пропускаем строки с метками конца таблиц
        if any(
            cell
            and str(cell).strip()
            in [
                "{end_table1}",
                "{end_table2}",
                "{end_table3}",
                "{{end_table1}}",
                "{{end_table2}}",
                "{{end_table3}}",
            ]
            for cell in row
        ):
            continue

        # Проверяем высоту листа
        current_height = calculate_current_height(current_sheet)
        if current_height + DEFAULT_ROW_HEIGHT > A4_HEIGHT_POINTS:
            sheet_number += 1
            current_sheet = current_sheet.parent.create_sheet(f"Лист{sheet_number}")
            set_sheet_margins(current_sheet)
            enforce_fit_to_page(current_sheet)
            current_row = 1

            # Копируем размеры столбцов
            copy_column_dimensions(template_sheet, current_sheet)

            # Создаем новую карту объединенных ячеек для нового листа
            sheet_merged_cells_map = current_sheet.merged_cells

        # Копируем форматирование строки
        copy_row_formatting(
            template_sheet, current_sheet, row_num, current_row, sheet_merged_cells_map
        )

        # Если это строка с меткой {executor} и есть исполнители
        if row_num == row_with_executor and executors and executor_column:
            # Обрабатываем первого исполнителя в текущей строке
            for col in range(1, template_sheet.max_column + 1):
                cell = current_sheet.cell(row=current_row, column=col)
                if cell.value:
                    if col == executor_column and "{executor}" in str(cell.value):
                        # Заменяем метку на первого исполнителя
                        cell.value = str(cell.value).replace("{executor}", executors[0])
                    else:
                        # Обрабатываем другие метки
                        processed_value = process_cell_markers(
                            protocol, str(cell.value)
                        )
                        if processed_value is not None:
                            cell.value = processed_value

            current_row += 1

            # Создаем дополнительные строки для остальных исполнителей
            for executor in executors[1:]:
                # Проверяем высоту листа перед добавлением новой строки
                current_height = calculate_current_height(current_sheet)
                if current_height + DEFAULT_ROW_HEIGHT > A4_HEIGHT_POINTS:
                    sheet_number += 1
                    current_sheet = current_sheet.parent.create_sheet(
                        f"Лист{sheet_number}"
                    )
                    set_sheet_margins(current_sheet)
                    enforce_fit_to_page(current_sheet)
                    current_row = 1

                    # Копируем размеры столбцов
                    copy_column_dimensions(template_sheet, current_sheet)

                    # Создаем новую карту объединенных ячеек для нового листа
                    sheet_merged_cells_map = current_sheet.merged_cells

                # Копируем форматирование строки для дополнительного исполнителя
                copy_row_formatting(
                    template_sheet,
                    current_sheet,
                    row_num,
                    current_row,
                    sheet_merged_cells_map,
                )

                # Заполняем значения в строке
                for col in range(1, template_sheet.max_column + 1):
                    cell = current_sheet.cell(row=current_row, column=col)
                    if cell.value:
                        if col == executor_column and "{executor}" in str(cell.value):
                            # Заменяем метку на текущего исполнителя
                            cell.value = str(cell.value).replace("{executor}", executor)
                        else:
                            # Для всех остальных ячеек ставим пустое значение
                            cell.value = ""

                current_row += 1
        else:
            # Обрабатываем обычные строки (без метки {executor})
            for col in range(1, template_sheet.max_column + 1):
                cell = current_sheet.cell(row=current_row, column=col)
                if cell.value:
                    # Обрабатываем метки
                    processed_value = process_cell_markers(protocol, str(cell.value))
                    if processed_value is not None:
                        cell.value = processed_value

            current_row += 1

    return current_sheet, current_row


def process_footer(
    protocol,
    template_sheet,
    current_sheet,
    footer_start,
    merged_cells_map,
    current_row,
    sheet_number,
):
    """
    Обрабатывает оставшиеся строки после последней таблицы (подвал протокола).
    """
    # Создаем новую карту объединенных ячеек для текущего листа
    sheet_merged_cells_map = current_sheet.merged_cells

    # Копируем и обрабатываем все оставшиеся строки
    for row_num in range(footer_start + 1, template_sheet.max_row + 1):
        # Пропускаем строки с метками концов таблиц
        row = list(
            template_sheet.iter_rows(min_row=row_num, max_row=row_num, values_only=True)
        )[0]
        if any(
            cell
            and str(cell).strip()
            in [
                "{end_table1}",
                "{end_table2}",
                "{end_table3}",
                "{{end_table1}}",
                "{{end_table2}}",
                "{{end_table3}}",
            ]
            for cell in row
        ):
            continue

        # Проверяем высоту листа
        current_height = calculate_current_height(current_sheet)
        if current_height + DEFAULT_ROW_HEIGHT > A4_HEIGHT_POINTS:
            sheet_number += 1
            current_sheet = current_sheet.parent.create_sheet(f"Лист{sheet_number}")
            set_sheet_margins(current_sheet)
            enforce_fit_to_page(current_sheet)
            current_row = 1

            # Копируем размеры столбцов
            copy_column_dimensions(template_sheet, current_sheet)

            # Создаем новую карту объединенных ячеек для нового листа
            sheet_merged_cells_map = current_sheet.merged_cells

        # Копируем размеры строк
        if row_num in template_sheet.row_dimensions:
            current_sheet.row_dimensions[current_row] = copy(
                template_sheet.row_dimensions[row_num]
            )

        # Копируем объединенные ячейки для текущей строки
        for merged_range in template_sheet.merged_cells.ranges:
            if merged_range.min_row == row_num:
                # Создаем новый диапазон с учетом смещения строк
                new_range = openpyxl.worksheet.cell_range.CellRange(
                    min_col=merged_range.min_col,
                    min_row=current_row,
                    max_col=merged_range.max_col,
                    max_row=current_row + (merged_range.max_row - merged_range.min_row),
                )
                sheet_merged_cells_map.add(new_range)

        # Копируем содержимое и стили ячеек, но без границ
        for col in range(1, template_sheet.max_column + 1):
            source_cell = template_sheet.cell(row=row_num, column=col)
            target_cell = current_sheet.cell(row=current_row, column=col)

            # Копируем значение
            target_cell.value = source_cell.value

            # Копируем стили кроме границ
            if source_cell.has_style:
                if source_cell.font:
                    target_cell.font = copy(source_cell.font)
                if source_cell.fill:
                    target_cell.fill = copy(source_cell.fill)
                if source_cell.alignment:
                    target_cell.alignment = copy(source_cell.alignment)
                if source_cell.number_format:
                    target_cell.number_format = source_cell.number_format
                if source_cell.protection:
                    target_cell.protection = copy(source_cell.protection)
                # Границы не копируем

        # Обрабатываем метки в строке
        for col in range(1, template_sheet.max_column + 1):
            cell = current_sheet.cell(row=current_row, column=col)
            if cell.value:
                # Обрабатываем метки
                processed_value = process_cell_markers(protocol, str(cell.value))
                if processed_value is not None:
                    cell.value = processed_value

        current_row += 1

    return current_sheet


def find_text_in_workbook(workbook, search_text):
    """
    Ищет указанный текст во всех листах книги Excel.
    Возвращает список всех найденных вхождений в формате [(лист, строка, столбец), ...].
    """
    try:
        results = []
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            for row_idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
                for col_idx, cell_value in enumerate(row, 1):
                    if (
                        cell_value
                        and isinstance(cell_value, str)
                        and search_text.lower() in cell_value.lower()
                    ):
                        results.append((sheet_name, row_idx, col_idx))
        return results
    except Exception as e:
        logger.error(f"Ошибка при поиске текста '{search_text}': {str(e)}")
        return []


def process_footer_test_protocol_number(protocol, text) -> _HeaderFooterPart:
    """
    Заменяет только {test_protocol_number} на значение из get_marker_value_title в тексте колонтитула.
    """
    # Достаём строку из объекта openpyxl
    orig_text = text
    if hasattr(text, "text"):
        text = text.text
    if not text or not isinstance(text, str):
        logger.info(f"Колонтитул: пустое или нестроковое значение: {text!r}")
        return orig_text
    value = get_marker_value_title(protocol, "test_protocol_number")
    result = text.replace("{test_protocol_number}", value)
    logger.info(f"Колонтитул: исходный: {text!r}, после замены: {result!r}")
    return _HeaderFooterPart(text=result)


def enforce_fit_to_page(sheet):
    """
    Явно выставляет fitToWidth=1, fitToHeight=0, scale=None для надёжности.
    """
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.page_setup.scale = 86
    logger.info(
        f"Применены параметры печати: fitToWidth={sheet.page_setup.fitToWidth}, "
        f"fitToHeight={sheet.page_setup.fitToHeight}, scale={sheet.page_setup.scale}"
    )


@api_view(["GET"])
@permission_classes([AllowAny])
def generate_protocol_excel(request):
    """
    Генерирует Excel файл протокола.
    """
    try:
        protocol_id = request.GET.get("protocol_id")
        if not protocol_id:
            return HttpResponseBadRequest("Не указан protocol_id")

        logger.info(f"Начинаем генерацию протокола {protocol_id}")

        protocol = Protocol.objects.get(id=protocol_id)

        if not protocol.excel_template:
            return HttpResponseBadRequest("У протокола отсутствует шаблон Excel")

        # Загружаем шаблон
        template_bytes = BytesIO(protocol.excel_template.file)
        template_workbook = openpyxl.load_workbook(template_bytes)
        template_sheet = template_workbook.active

        logger.info("Шаблон успешно загружен")
        logger.debug(
            f"Размеры шаблона: строк={template_sheet.max_row}, столбцов={template_sheet.max_column}"
        )

        # Создаем новый файл
        new_workbook = openpyxl.Workbook()
        new_sheet = new_workbook.active
        new_sheet.title = "Лист 1"
        set_sheet_margins(new_sheet)
        enforce_fit_to_page(new_sheet)

        # Копирование нижнего колонтитула с обработкой только test_protocol_number
        if hasattr(template_sheet, "oddFooter") and hasattr(new_sheet, "oddFooter"):
            new_sheet.oddFooter.left = process_footer_test_protocol_number(
                protocol, template_sheet.oddFooter.left
            )
            new_sheet.oddFooter.center = process_footer_test_protocol_number(
                protocol, template_sheet.oddFooter.center
            )
            new_sheet.oddFooter.right = process_footer_test_protocol_number(
                protocol, template_sheet.oddFooter.right
            )
        if hasattr(template_sheet, "evenFooter") and hasattr(new_sheet, "evenFooter"):
            new_sheet.evenFooter.left = process_footer_test_protocol_number(
                protocol, template_sheet.evenFooter.left
            )
            new_sheet.evenFooter.center = process_footer_test_protocol_number(
                protocol, template_sheet.evenFooter.center
            )
            new_sheet.evenFooter.right = process_footer_test_protocol_number(
                protocol, template_sheet.evenFooter.right
            )
        if hasattr(template_sheet, "firstFooter") and hasattr(new_sheet, "firstFooter"):
            new_sheet.firstFooter.left = process_footer_test_protocol_number(
                protocol, template_sheet.firstFooter.left
            )
            new_sheet.firstFooter.center = process_footer_test_protocol_number(
                protocol, template_sheet.firstFooter.center
            )
            new_sheet.firstFooter.right = process_footer_test_protocol_number(
                protocol, template_sheet.firstFooter.right
            )

        copy_column_dimensions(template_sheet, new_sheet)

        # Устанавливаем размеры листа как в шаблоне
        new_sheet._current_row = template_sheet.max_row
        new_sheet._max_column = template_sheet.max_column

        # Создаем карту объединенных ячеек
        merged_cells_map = new_sheet.merged_cells

        # Обрабатываем шапку
        header_end = process_header(
            protocol, template_sheet, new_sheet, merged_cells_map
        )
        if header_end == 0:
            return HttpResponseBadRequest("Ошибка при обработке шапки протокола")

        # Обрабатываем заголовок и условия отбора
        table_start = process_header_and_conditions(
            protocol, template_sheet, new_sheet, header_end, merged_cells_map
        )

        # Обрабатываем первую таблицу с методами
        current_sheet = process_methods_table(
            protocol,
            template_sheet,
            new_sheet,
            table_start,
            merged_cells_map,
            new_sheet.max_row + 1,
            1,
        )
        if not current_sheet:
            return HttpResponseBadRequest("Ошибка при обработке таблицы методов")

        # Находим конец первой таблицы
        table1_end = None
        for row_num in range(table_start, template_sheet.max_row + 1):
            row = list(
                template_sheet.iter_rows(
                    min_row=row_num, max_row=row_num, values_only=True
                )
            )[0]
            if any(cell and str(cell).strip() == "{end_table1}" for cell in row):
                table1_end = row_num
                break

        if table1_end:
            # Обрабатываем данные между таблицами 1 и 2
            current_sheet, current_row = process_between_tables(
                protocol,
                template_sheet,
                current_sheet,
                table1_end,
                merged_cells_map,
                current_sheet.max_row + 1,
                len(new_workbook.sheetnames),
            )

            # Обрабатываем вторую таблицу
            current_sheet = process_equipment_table(
                protocol,
                template_sheet,
                current_sheet,
                table1_end + 1,
                merged_cells_map,
                    current_row,
                len(new_workbook.sheetnames),
            )
            if not current_sheet:
                return HttpResponseBadRequest(
                    "Ошибка при обработке таблицы оборудования"
                )

            # Находим конец второй таблицы
            table2_end = None
            for row_num in range(table1_end + 1, template_sheet.max_row + 1):
                row = list(
                    template_sheet.iter_rows(
                        min_row=row_num, max_row=row_num, values_only=True
                    )
                )[0]
                if any(cell and str(cell).strip() == "{end_table2}" for cell in row):
                    table2_end = row_num
                    break

            if table2_end:
                # Обрабатываем данные между таблицами 2 и 3
                current_sheet, current_row = process_between_tables(
                    protocol,
                    template_sheet,
                    current_sheet,
                    table2_end,
                    merged_cells_map,
                    current_sheet.max_row + 1,
                    len(new_workbook.sheetnames),
                )

                # Обрабатываем третью таблицу
                current_sheet = process_nd_table(
                    protocol,
                    template_sheet,
                    current_sheet,
                    table2_end + 1,
                    merged_cells_map,
                    current_row,
                    len(new_workbook.sheetnames),
                )
                if not current_sheet:
                    return HttpResponseBadRequest("Ошибка при обработке таблицы НД")

                # Находим конец третьей таблицы
                table3_end = None
                for row_num in range(table2_end + 1, template_sheet.max_row + 1):
                    row = list(
                        template_sheet.iter_rows(
                            min_row=row_num, max_row=row_num, values_only=True
                        )
                    )[0]
                    if any(
                        cell and str(cell).strip() == "{end_table3}" for cell in row
                    ):
                        table3_end = row_num
                        break

                if table3_end:
                    # Обрабатываем оставшиеся строки после таблицы 3
                    current_sheet = process_footer(
                        protocol,
                    template_sheet,
                        current_sheet,
                        table3_end,
                    merged_cells_map,
                        current_sheet.max_row + 1,
                        len(new_workbook.sheetnames),
                    )

        logger.info("Начинаем сохранение файла")
        logger.debug(
            f"Размеры нового листа: строк={new_sheet.max_row}, столбцов={new_sheet.max_column}"
        )

        try:
            # Ищем все ячейки с текстом "конец протокола"
            end_protocol_locations = find_text_in_workbook(
                new_workbook, "конец протокола"
            )
            if end_protocol_locations:
                for location in end_protocol_locations:
                    sheet_name, row_idx, col_idx = location
                    logger.info(
                        f"Найден текст 'конец протокола' в листе '{sheet_name}', ячейка: строка {row_idx}, столбец {col_idx}"
                    )
                logger.info(
                    f"Всего найдено вхождений 'конец протокола': {len(end_protocol_locations)}"
                )
            else:
                logger.info(
                    "Текст 'конец протокола' не найден в сгенерированном документе"
                )

            # Сохраняем результат
        output = BytesIO()
        new_workbook.save(output)
        output.seek(0)

            logger.info("Файл успешно сохранен в BytesIO")

        response = HttpResponse(
                output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
            response["Content-Disposition"] = (
                f'attachment; filename="protocol_{protocol_id}.xlsx"'
            )

            logger.info("Протокол успешно сгенерирован")
        return response

        except Exception as save_error:
            logger.error(f"Ошибка при сохранении файла: {str(save_error)}")
            return HttpResponseBadRequest("Ошибка при сохранении файла")

    except Protocol.DoesNotExist:
        logger.error(f"Протокол с id {protocol_id} не найден")
        return HttpResponseBadRequest("Протокол не найден")
    except Exception as e:
        logger.error(f"Ошибка при генерации протокола: {str(e)}")
        return HttpResponseBadRequest(f"Ошибка при генерации протокола: {str(e)}")
