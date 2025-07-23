from datetime import datetime
from calendar import monthrange
from openpyxl import load_workbook
from openpyxl.worksheet.cell_range import CellRange
from copy import copy
from io import BytesIO
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import AllowAny
from rest_framework.response import Response
from django.http import HttpResponse
from ..models import ReportTemplate, Sample, Calculation
from ..utils.protocol_generator_utils import (
    copy_cell_style,
    copy_row_formatting,
    copy_column_dimensions,
)
from ..utils.employee_utils import get_employee_name


def find_month_markers(sheet):
    """Находит метки месяцев в шаблоне и их строки"""
    month_markers = {}
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                if "{start_" in cell.value and "_january}" in cell.value:
                    month = cell.value.split("start_")[1].split("}")[0]
                    month_markers[month] = {"start": cell.row, "end": None}
                elif "{end_" in cell.value and "_january}" in cell.value:
                    month = cell.value.split("end_")[1].split("}")[0]
                    if month in month_markers:
                        month_markers[month]["end"] = cell.row
    return month_markers


def find_template_row(sheet, start_row, end_row):
    """Находит шаблонную строку между start_row и end_row"""
    for row_idx in range(start_row, end_row + 1):
        for cell in sheet[row_idx]:
            if cell.value and isinstance(cell.value, str):
                if any(
                    marker in cell.value
                    for marker in ["{id}", "{reg_number}", "{sampling_location}"]
                ):
                    return row_idx
    return None


def get_month_data(samples, year, month, global_start_date, global_end_date):
    """Получает данные для конкретного месяца с учетом глобального диапазона дат"""
    start_date = datetime(year, month, 1)
    _, last_day = monthrange(year, month)
    end_date = datetime(year, month, last_day)

    if start_date < global_start_date:
        start_date = global_start_date
    if end_date > global_end_date:
        end_date = global_end_date

    return samples.filter(
        sampling_date__gte=start_date, sampling_date__lte=end_date
    ).order_by("sampling_date")


def fill_template_row(row, sample, template_row, sheet):
    """Заполняет строку данными из sample"""
    # Находим все объединенные диапазоны для текущей строки
    merged_ranges = []
    for merged_range in sheet.merged_cells.ranges:
        if merged_range.min_row <= row <= merged_range.max_row:
            merged_ranges.append(merged_range)

    for cell in template_row:
        # Проверяем, является ли ячейка частью объединенного диапазона
        is_merged = False
        master_cell = None
        current_range = None
        for merged_range in merged_ranges:
            if (
                merged_range.min_col <= cell.column <= merged_range.max_col
                and merged_range.min_row <= row <= merged_range.max_row
            ):
                is_merged = True
                current_range = merged_range
                # Получаем главную ячейку объединенного диапазона
                master_cell = sheet.cell(
                    row=merged_range.min_row, column=merged_range.min_col
                )
                break

        # Используем значение из главной ячейки для объединенных ячеек
        cell_value = None
        if is_merged and master_cell:
            if cell.column == master_cell.column and row == master_cell.row:
                cell_value = master_cell.value
        else:
            cell_value = cell.value

        if not cell_value or not isinstance(cell_value, str):
            continue

        new_cell = sheet.cell(row=row, column=cell.column)

        value = cell_value
        if "{id}" in value:
            new_cell.value = value.replace("{id}", str(sample.id))
        elif "{reg_number}" in value:
            new_cell.value = value.replace("{reg_number}", sample.registration_number)
        elif "{sampling_location}" in value:
            new_cell.value = value.replace(
                "{sampling_location}", sample.sampling_location_detail or "-"
            )
        elif "{sampling_date}" in value:
            date_str = (
                sample.sampling_date.strftime("%d.%m.%Y")
                if sample.sampling_date
                else "-"
            )
            new_cell.value = value.replace("{sampling_date}", date_str)
        elif "{receiving_date}" in value:
            date_str = (
                sample.receiving_date.strftime("%d.%m.%Y")
                if sample.receiving_date
                else "-"
            )
            new_cell.value = value.replace("{receiving_date}", date_str)
        elif "{conditions}" in value:
            conditions = sample.protocol.selection_conditions if sample.protocol else {}
            conditions_str = (
                ", ".join(f"{k}: {v}" for k, v in conditions.items())
                if conditions
                else "-"
            )
            new_cell.value = value.replace("{conditions}", conditions_str)
        elif "{quantity}" in value:
            quantity = sample.calculations.filter(is_deleted=False).count()
            new_cell.value = value.replace("{quantity}", str(quantity))
        elif "{executor}" in value:
            executors = {
                get_employee_name(calc.executor)
                for calc in sample.calculations.filter(is_deleted=False)
                if calc.executor
            }
            executors_str = ", ".join(sorted(executors)) if executors else "-"
            new_cell.value = value.replace("{executor}", executors_str)
        elif "{protocol}" in value:
            protocol_num = (
                sample.protocol.test_protocol_number if sample.protocol else "-"
            )
            new_cell.value = value.replace("{protocol}", protocol_num)
        else:
            if is_merged and master_cell:
                if cell.column == master_cell.column and row == master_cell.row:
                    new_cell.value = value
                else:
                    new_cell.value = None
            else:
                new_cell.value = value

        # Копируем стили после установки значения
        try:
            if is_merged and master_cell:
                copy_cell_style(master_cell, new_cell)
            else:
                copy_cell_style(cell, new_cell)
        except AttributeError:
            # Пропускаем копирование стилей, если возникла ошибка
            pass


def copy_row_with_merged_cells(
    source_sheet, target_sheet, source_row, target_row, merged_cells_map
):
    """
    Копирует строку с сохранением форматирования и корректной обработкой объединенных ячеек.
    """
    # Копируем размеры строк
    if source_row in source_sheet.row_dimensions:
        target_sheet.row_dimensions[target_row] = copy(
            source_sheet.row_dimensions[source_row]
        )

    # Находим все объединенные диапазоны для текущей строки
    merged_ranges = []
    for merged_range in source_sheet.merged_cells.ranges:
        if merged_range.min_row <= source_row <= merged_range.max_row:
            merged_ranges.append(merged_range)
            # Создаем новый диапазон с учетом смещения строк
            if merged_range.min_row == source_row:
                new_range = CellRange(
                    min_col=merged_range.min_col,
                    min_row=target_row,
                    max_col=merged_range.max_col,
                    max_row=target_row + (merged_range.max_row - merged_range.min_row),
                )
                merged_cells_map.add(new_range)

    # Копируем содержимое и стили ячеек
    for col in range(1, source_sheet.max_column + 1):
        source_cell = source_sheet.cell(row=source_row, column=col)
        target_cell = target_sheet.cell(row=target_row, column=col)

        # Проверяем, является ли текущая ячейка частью объединенного диапазона
        is_merged = False
        master_cell = None
        for merged_range in merged_ranges:
            if (
                merged_range.min_col <= col <= merged_range.max_col
                and merged_range.min_row <= source_row <= merged_range.max_row
            ):
                is_merged = True
                # Получаем главную ячейку объединенного диапазона
                master_cell = source_sheet.cell(
                    row=merged_range.min_row, column=merged_range.min_col
                )
                break

        # Копируем значение из главной ячейки, если это объединенная ячейка
        if is_merged and master_cell:
            if col == merged_range.min_col and source_row == merged_range.min_row:
                target_cell.value = master_cell.value
            else:
                target_cell.value = None
        else:
            target_cell.value = source_cell.value

        # Копируем стили
        try:
            if is_merged and master_cell:
                copy_cell_style(master_cell, target_cell)
            else:
                copy_cell_style(source_cell, target_cell)
        except AttributeError:
            # Пропускаем копирование стилей, если возникла ошибка
            pass


@api_view(["POST"])
@permission_classes([AllowAny])
def generate_report(request):
    """Генерирует отчет Excel на основе шаблона."""
    try:
        laboratory_id = request.data.get("laboratory_id")
        department_id = request.data.get("department_id")
        start_date = datetime.strptime(request.data.get("start_date"), "%Y-%m-%d")
        end_date = datetime.strptime(request.data.get("end_date"), "%Y-%m-%d")

        if not laboratory_id or not start_date or not end_date:
            return Response({"detail": "Не указаны обязательные параметры"}, status=400)

        # Получаем активный шаблон
        template_query = ReportTemplate.objects.filter(
            laboratory_id=laboratory_id, is_active=True
        )
        if department_id:
            template_query = template_query.filter(department_id=department_id)
        template = template_query.first()

        if not template:
            return Response({"detail": "Активный шаблон не найден"}, status=400)

        wb = load_workbook(BytesIO(template.file))
        sheet = wb.active

        new_sheet = wb.create_sheet("Результат")
        wb.active = new_sheet

        copy_column_dimensions(sheet, new_sheet)

        # Получаем все пробы за период
        samples = Sample.objects.filter(
            laboratory_id=laboratory_id,
            sampling_date__gte=start_date,
            sampling_date__lte=end_date,
            is_deleted=False,
        ).order_by("sampling_date")

        if department_id:
            samples = samples.filter(department_id=department_id)

        if not samples.exists():
            return Response({"detail": "Нет данных за выбранный период"}, status=400)

        # Копируем заголовок таблицы
        current_row = 1
        for row in range(1, sheet.max_row + 1):
            row_values = list(
                sheet.iter_rows(min_row=row, max_row=row, values_only=True)
            )[0]
            # Если встретили первый месяц, останавливаемся
            if any(
                isinstance(cell, str) and "ЯНВАРЬ" in str(cell) for cell in row_values
            ):
                break
            copy_row_formatting(sheet, new_sheet, row, current_row)
            current_row += 1

        # Обрабатываем каждый месяц
        months = {
            1: ("ЯНВАРЬ", "january"),
            2: ("ФЕВРАЛЬ", "february"),
            3: ("МАРТ", "march"),
            4: ("АПРЕЛЬ", "april"),
            5: ("МАЙ", "may"),
            6: ("ИЮНЬ", "june"),
            7: ("ИЮЛЬ", "july"),
            8: ("АВГУСТ", "august"),
            9: ("СЕНТЯБРЬ", "september"),
            10: ("ОКТЯБРЬ", "october"),
            11: ("НОЯБРЬ", "november"),
            12: ("ДЕКАБРЬ", "december"),
        }

        # Находим шаблонную строку и метки для каждого месяца
        month_templates = {}
        current_month = None
        for row in range(1, sheet.max_row + 1):
            row_values = list(
                sheet.iter_rows(min_row=row, max_row=row, values_only=True)
            )[0]
            row_str = str(row_values)

            # Определяем текущий месяц
            for month_num, (month_name, month_code) in months.items():
                if month_name in row_str:
                    current_month = month_num
                    month_templates[month_num] = {
                        "name": month_name,
                        "code": month_code,
                    }
                    break

            if current_month and current_month in month_templates:
                if "{start_" + month_templates[current_month]["code"] + "}" in row_str:
                    month_templates[current_month]["start_row"] = row
                elif "{end_" + month_templates[current_month]["code"] + "}" in row_str:
                    month_templates[current_month]["end_row"] = row
                elif all(marker in row_str for marker in ["{id}", "{reg_number}"]):
                    month_templates[current_month]["template_row"] = row
                    month_templates[current_month]["template_cells"] = list(sheet[row])

        # Группируем пробы по месяцам
        start_year = start_date.year
        end_year = end_date.year

        for year in range(start_year, end_year + 1):
            for month_num in range(1, 13):
                if month_num not in month_templates:
                    continue

                # Получаем данные месяца
                month_samples = samples.filter(
                    sampling_date__year=year, sampling_date__month=month_num
                )

                # Если в месяце нет данных, пропускаем его
                if not month_samples.exists():
                    continue

                # Копируем название месяца
                month_name_row = None
                for row in range(1, sheet.max_row + 1):
                    row_values = list(
                        sheet.iter_rows(min_row=row, max_row=row, values_only=True)
                    )[0]
                    if any(
                        isinstance(cell, str) and months[month_num][0] in str(cell)
                        for cell in row_values
                    ):
                        month_name_row = row
                        break

                if month_name_row:
                    copy_row_formatting(sheet, new_sheet, month_name_row, current_row)
                    current_row += 1

                # Заполняем данные месяца
                template = month_templates[month_num]
                template_cells = template["template_cells"]

                for idx, sample in enumerate(month_samples, 1):
                    copy_row_formatting(
                        sheet, new_sheet, template["template_row"], current_row
                    )

                    # Получаем данные для строки
                    calculations = sample.calculations.filter(is_deleted=False)
                    executors = {
                        get_employee_name(calc.executor)
                        for calc in sample.calculations.filter(is_deleted=False)
                        if calc.executor
                    }
                    conditions = (
                        sample.protocol.selection_conditions if sample.protocol else {}
                    )
                    conditions_str = (
                        ", ".join(f"{k}: {v}" for k, v in conditions.items())
                        if conditions
                        else "-"
                    )

                    # Заполняем ячейки
                    for cell in template_cells:
                        if not cell.value or not isinstance(cell.value, str):
                            continue

                        new_cell = new_sheet.cell(row=current_row, column=cell.column)
                        value = str(cell.value)

                        if "{id}" in value:
                            new_cell.value = str(idx)
                        elif "{reg_number}" in value:
                            new_cell.value = sample.registration_number
                        elif "{sampling_location}" in value:
                            new_cell.value = sample.sampling_location_detail or "-"
                        elif "{sampling_date}" in value:
                            new_cell.value = (
                                sample.sampling_date.strftime("%d.%m.%Y")
                                if sample.sampling_date
                                else "-"
                            )
                        elif "{receiving_date}" in value:
                            new_cell.value = (
                                sample.receiving_date.strftime("%d.%m.%Y")
                                if sample.receiving_date
                                else "-"
                            )
                        elif "{conditions}" in value:
                            new_cell.value = conditions_str
                        elif "{quantity}" in value:
                            new_cell.value = str(calculations.count())
                        elif "{executor}" in value:
                            new_cell.value = (
                                ", ".join(sorted(executors)) if executors else "-"
                            )
                        elif "{protocol}" in value:
                            new_cell.value = (
                                sample.protocol.test_protocol_number
                                if sample.protocol
                                else "-"
                            )

                    current_row += 1

        # Удаляем исходный лист
        wb.remove(sheet)

        # Сохраняем результат
        output = BytesIO()
        wb.save(output)

        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="report.xlsx"'
        return response

    except ValueError as e:
        return Response({"detail": str(e)}, status=400)
    except Exception as e:
        return Response(
            {"detail": "Произошла ошибка при формировании отчета"}, status=500
        )
