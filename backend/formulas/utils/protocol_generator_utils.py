import logging
from copy import copy
import openpyxl
from django.utils import formats

logger = logging.getLogger(__name__)

# Константы для размеров листа А4
A4_HEIGHT_POINTS = 841.89  # Высота А4 в точках
DEFAULT_ROW_HEIGHT = 15  # Стандартная высота строки в Excel


def format_decimal_ru(value) -> str:
    """
    Форматирует десятичное число для отображения в русском формате.
    """
    if value is None:
        return ""

    try:
        if isinstance(value, (int, float)):
            return str(value).replace(".", ",")
        return str(value)
    except:
        return str(value)


def copy_cell_style(source_cell, target_cell):
    """
    Безопасное копирование стилей из одной ячейки в другую
    """
    if not source_cell or not source_cell.has_style:
        return

    target_cell.font = copy(source_cell.font)
    target_cell.fill = copy(source_cell.fill)
    target_cell.border = copy(source_cell.border)
    target_cell.alignment = copy(source_cell.alignment)
    target_cell.number_format = source_cell.number_format
    target_cell.protection = copy(source_cell.protection)


def copy_row_with_styles(
    source_sheet: openpyxl.worksheet.worksheet.Worksheet,
    target_sheet: openpyxl.worksheet.worksheet.Worksheet,
    source_row: int,
    target_row: int,
) -> None:
    """
    Копирует строку с сохранением стилей из исходного листа в целевой.
    """
    try:
        logger.debug(f"Начинаем копирование строки {source_row} в строку {target_row}")

        # Получаем максимальное количество столбцов
        max_col = source_sheet.max_column
        logger.debug(f"Максимальное количество столбцов: {max_col}")

        # Копируем каждую ячейку в строке
        for col in range(1, max_col + 1):
            try:
                source_cell = source_sheet.cell(row=source_row, column=col)
                target_cell = target_sheet.cell(row=target_row, column=col)

                target_cell.value = source_cell.value
                copy_cell_style(source_cell, target_cell)

                logger.debug(
                    f"Скопирована ячейка [{source_row}, {col}] -> [{target_row}, {col}]"
                )

            except Exception as cell_error:
                logger.error(
                    f"Ошибка при копировании ячейки [{source_row}, {col}]: {str(cell_error)}"
                )
                continue

        logger.debug(f"Успешно скопирована строка {source_row} -> {target_row}")

    except Exception as e:
        logger.error(f"Ошибка при копировании строки {source_row}: {str(e)}")
        raise


def copy_row_formatting(
    source_sheet, target_sheet, source_row, target_row, merged_cells_map=None
):
    """
    Копирует все форматирование строки: стили, размеры и объединенные ячейки
    """
    if source_row in source_sheet.row_dimensions:
        target_sheet.row_dimensions[target_row] = copy(
            source_sheet.row_dimensions[source_row]
        )

    copy_row_with_styles(source_sheet, target_sheet, source_row, target_row)

    if merged_cells_map is not None:
        for merged_range in source_sheet.merged_cells.ranges:
            if merged_range.min_row == source_row:
                new_range = openpyxl.worksheet.cell_range.CellRange(
                    min_col=merged_range.min_col,
                    min_row=target_row,
                    max_col=merged_range.max_col,
                    max_row=target_row + (merged_range.max_row - merged_range.min_row),
                )
                merged_cells_map.add(new_range)


def calculate_current_height(sheet) -> float:
    """
    Рассчитывает текущую высоту листа в точках
    """
    total_height = 0
    for row_num in range(1, sheet.max_row + 1):
        if row_num in sheet.row_dimensions:
            total_height += sheet.row_dimensions[row_num].height or DEFAULT_ROW_HEIGHT
        else:
            total_height += DEFAULT_ROW_HEIGHT
    return total_height


def calculate_header_height(
    template_sheet, table_header_start, table_header_end
) -> float:
    """
    Рассчитывает высоту шапки таблицы в точках
    """
    header_height = 0
    for row_num in range(table_header_start + 1, table_header_end):
        if row_num in template_sheet.row_dimensions:
            header_height += (
                template_sheet.row_dimensions[row_num].height or DEFAULT_ROW_HEIGHT
            )
        else:
            header_height += DEFAULT_ROW_HEIGHT
    return header_height


def copy_column_dimensions(source_sheet, target_sheet):
    """
    Копирует размеры столбцов из исходного листа в целевой.
    """
    try:
        for key, value in source_sheet.column_dimensions.items():
            target_sheet.column_dimensions[key].width = value.width
            target_sheet.column_dimensions[key].hidden = value.hidden
    except Exception as e:
        logger.error(f"Ошибка при копировании размеров столбцов: {str(e)}")
