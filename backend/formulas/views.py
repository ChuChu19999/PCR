import os
from django.conf import settings
from django.db import connections
from django.utils import timezone
from django.shortcuts import get_object_or_404
from rest_framework import viewsets, status
from rest_framework.decorators import api_view, permission_classes, action
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework.response import Response
from openpyxl import load_workbook
from openpyxl.styles import Font
from .models import Laboratory, Department, ResearchMethod, ResearchObject
from .serializers import (
    UserSerializer,
    LaboratorySerializer,
    DepartmentSerializer,
    ResearchMethodSerializer,
    ResearchObjectSerializer,
)
import logging
from django.db.models import Q
from decimal import Decimal, ROUND_HALF_UP


logger = logging.getLogger(__name__)


def get_user_role(hach_snils):
    with connections["access_control"].cursor() as cursor:
        cursor.execute(
            """
            SELECT r.name
            FROM access_control.users_roles_systems urs
            JOIN access_control.roles r ON urs.role_id = r.id
            WHERE urs.user_hash=%s
            AND urs.is_active=TRUE
            AND urs.system_id = 8
            LIMIT 1;
        """,
            [hach_snils],
        )
        row = cursor.fetchone()
        return row[0] if row else None


class UserViewSet(viewsets.ReadOnlyModelViewSet):
    permission_classes = [IsAuthenticated]
    serializer_class = UserSerializer

    def list(self, request, *args, **kwargs):
        decoded_token = getattr(request, "decoded_token", None)
        if decoded_token is None:
            return Response({"error": "Invalid token"}, status=401)

        hash_snils = decoded_token.get("hashSnils")
        if not hash_snils:
            return Response({"error": "Hashsnils not found in token"}, status=400)

        user_role = get_user_role(hash_snils)

        is_staff = user_role == "editor"

        if not user_role:
            return Response({"error": "Access denied"}, status=403)

        response_data = {
            "personnelNumber": decoded_token.get("personnelNumber"),
            "departmentNumber": decoded_token.get("departmentNumber"),
            "fullName": decoded_token.get("fullName"),
            "preferred_username": decoded_token.get("preferred_username"),
            "email": decoded_token.get("email"),
            "hashSnils": decoded_token.get("hashSnils"),
            "is_staff": is_staff,
        }
        serializer = self.get_serializer(data=response_data)
        serializer.is_valid(raise_exception=True)
        return Response(serializer.data)

    def retrieve(self, request, *args, **kwargs):
        return self.list(request, *args, **kwargs)


@permission_classes([AllowAny])
class LaboratoryViewSet(viewsets.ModelViewSet):
    serializer_class = LaboratorySerializer

    def get_queryset(self):
        return Laboratory.objects.filter(is_deleted=False)

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        # Помечаем как удаленные все связанные research-pages
        ResearchObject.objects.filter(laboratory=instance).update(
            is_deleted=True, deleted_at=timezone.now()
        )
        instance.mark_as_deleted()
        return Response(status=status.HTTP_204_NO_CONTENT)


@permission_classes([AllowAny])
class DepartmentViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated]
    serializer_class = DepartmentSerializer

    def get_queryset(self):
        return Department.objects.filter(is_deleted=False).select_related("laboratory")

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        # Помечаем как удаленные все связанные research-pages
        ResearchObject.objects.filter(department=instance).update(
            is_deleted=True, deleted_at=timezone.now()
        )
        instance.is_deleted = True
        instance.deleted_at = timezone.now()
        instance.save()
        return Response(status=status.HTTP_204_NO_CONTENT)

    def create(self, request, *args, **kwargs):
        laboratory_id = request.data.get("laboratory")
        get_object_or_404(Laboratory, id=laboratory_id)
        return super().create(request, *args, **kwargs)


@api_view(["POST"])
@permission_classes([AllowAny])
def save_excel(request):
    try:
        data = request.data.get("data")
        styles = request.data.get("styles", {})

        logger.info(f"Получены данные для сохранения: {data}")
        logger.info(f"Получены стили для сохранения: {styles}")

        if not data:
            logger.error("Данные не предоставлены")
            return Response(
                {"error": "Данные не предоставлены"}, status=status.HTTP_400_BAD_REQUEST
            )

        file_path = os.path.join(settings.MEDIA_ROOT, "shablon.xlsx")
        logger.info(f"Путь к файлу Excel: {file_path}")

        if not os.path.exists(file_path):
            logger.error(f"Файл не найден: {file_path}")
            return Response(
                {"error": "Файл шаблона не найден"}, status=status.HTTP_404_NOT_FOUND
            )

        workbook = load_workbook(file_path, data_only=False)
        worksheet = workbook.active

        for row_idx, row_data in enumerate(data):
            if row_idx < 8:
                cell = worksheet.cell(row=row_idx + 1, column=1, value=row_data[0])
                logger.info(
                    f"Обновлена ячейка [{row_idx + 1}, 1] значением: {row_data[0]}"
                )

                # Применяем стили, если они есть для данной ячейки
                cell_key = f"{row_idx}-0"
                if cell_key in styles:
                    style = styles[cell_key]
                    logger.info(f"Применение стилей для ячейки {cell_key}: {style}")

                    # Получаем размер шрифта и конвертируем его в целое число
                    font_size = style.get("fontSize", "14")
                    if isinstance(font_size, str):
                        # Удаляем 'px' и конвертируем в float, затем в int
                        font_size = font_size.replace("px", "")
                        try:
                            font_size = int(float(font_size))
                        except (ValueError, TypeError):
                            font_size = 14

                    font = Font(
                        name="Times New Roman",
                        bold=style.get("fontWeight") == "bold",
                        italic=style.get("fontStyle") == "italic",
                        size=font_size,
                    )
                    cell.font = font

                    logger.info(
                        f"Стили применены: bold={font.bold}, italic={font.italic}, size={font.size}"
                    )

        logger.info("Сохранение файла...")
        workbook.save(file_path)
        logger.info(f"Excel файл успешно обновлен со стилями")

        return Response(
            {"message": "Файл успешно обновлен", "applied_styles": styles},
            status=status.HTTP_200_OK,
        )
    except Exception as e:
        logger.error(f"Ошибка при сохранении Excel файла: {str(e)}", exc_info=True)
        return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(["GET"])
@permission_classes([AllowAny])
def get_excel_styles(request):
    try:
        file_path = os.path.join(settings.MEDIA_ROOT, "shablon.xlsx")
        logger.info(f"Получение стилей из файла: {file_path}")

        workbook = load_workbook(file_path, data_only=False)
        worksheet = workbook.active

        styles = {}

        # Получаем стили для первых 8 строк первой колонки
        for row_idx in range(8):
            cell = worksheet.cell(row=row_idx + 1, column=1)
            cell_key = f"{row_idx}-0"

            if cell.font:
                style = {}
                if cell.font.bold:
                    style["fontWeight"] = "bold"
                    logger.info(f"Ячейка {cell_key}: установлен жирный шрифт")
                if cell.font.italic:
                    style["fontStyle"] = "italic"
                    logger.info(f"Ячейка {cell_key}: установлен курсив")
                if cell.font.size:
                    style["fontSize"] = f"{cell.font.size}px"
                    logger.info(
                        f"Ячейка {cell_key}: установлен размер шрифта {cell.font.size}px"
                    )

                if style:
                    styles[cell_key] = style
                    logger.info(f"Добавлены стили для ячейки {cell_key}: {style}")

        logger.info(f"Всего найдено стилей: {len(styles)}")
        logger.info(f"Возвращаемые стили: {styles}")
        return Response({"styles": styles}, status=status.HTTP_200_OK)
    except Exception as e:
        logger.error(f"Ошибка при получении стилей Excel: {str(e)}", exc_info=True)
        return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(["GET"])
@permission_classes([AllowAny])
def check_status_api(request):
    try:
        laboratories = LaboratoryViewSet()
        laboratories.request = request
        laboratories.format_kwarg = None

        response = laboratories.list(request)

        return Response(
            {"status": "online"},
            status=status.HTTP_200_OK,
        )

    except Exception as e:
        return Response(
            {"status": "offline"},
            status=status.HTTP_503_SERVICE_UNAVAILABLE,
        )


class ResearchMethodViewSet(viewsets.ModelViewSet):
    serializer_class = ResearchMethodSerializer
    permission_classes = [AllowAny]

    def get_queryset(self):
        queryset = ResearchMethod.objects.all()

        if self.action == "list":
            queryset = queryset.filter(is_active=True, is_deleted=False)

        search = self.request.query_params.get("search", None)
        if search:
            queryset = queryset.filter(
                Q(name__icontains=search)
                | Q(nd_code__icontains=search)
                | Q(nd_name__icontains=search)
            )

        rounding_type = self.request.query_params.get("rounding_type", None)
        if rounding_type:
            queryset = queryset.filter(rounding_type=rounding_type)

        return queryset.order_by("name")

    def perform_create(self, serializer):
        serializer.save()

    def update(self, request, *args, **kwargs):
        """
        Обновляет метод исследования.
        Для convergence_conditions ожидает формат:
        {
            "convergence_conditions": {
                "formulas": ["формула1", "формула2"]
            }
        }
        """
        instance = self.get_object()

        # Проверяем, есть ли convergence_conditions в запросе
        if "convergence_conditions" in request.data:
            conditions = request.data.get("convergence_conditions")
            if not isinstance(conditions, dict) or "formulas" not in conditions:
                return Response(
                    {"error": "Неверный формат условий сходимости"},
                    status=status.HTTP_400_BAD_REQUEST,
                )

        serializer = self.get_serializer(instance, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)

        return Response(serializer.data)

    def perform_update(self, serializer):
        serializer.save()

    @action(detail=True, methods=["post"])
    def toggle_active(self, request, pk=None):
        research_method = self.get_object()
        research_method.is_active = not research_method.is_active
        research_method.save()

        return Response({"status": "success", "is_active": research_method.is_active})

    @action(detail=True, methods=["post"])
    def mark_deleted(self, request, pk=None):
        research_method = self.get_object()
        research_method.is_deleted = True
        research_method.deleted_at = timezone.now()
        research_method.save()

        return Response(
            {"status": "success", "message": "Метод исследования помечен как удаленный"}
        )

    @action(detail=False, methods=["get"])
    def active(self, request):
        queryset = self.get_queryset().filter(is_active=True, is_deleted=False)
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=["get"])
    def deleted(self, request):
        queryset = self.get_queryset().filter(is_deleted=True)
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)


class ResearchPageViewSet(viewsets.ModelViewSet):
    serializer_class = ResearchObjectSerializer
    permission_classes = [AllowAny]

    def get_queryset(self):
        return ResearchObject.objects.filter(is_deleted=False)

    def retrieve(self, request, *args, **kwargs):
        instance = self.get_object()
        serializer = self.get_serializer(instance)
        return Response(serializer.data)

    @action(detail=True, methods=["get"])
    def current_method(self, request, pk=None):
        """
        Получает полную информацию о текущем методе исследования.
        """
        try:
            instance = self.get_object()

            current_method_id = request.query_params.get("current_method_id")
            if not current_method_id:
                return Response(
                    {"error": "Не указан ID текущего метода исследования"},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            # Проверяем, что метод принадлежит этой странице
            if not instance.research_methods.filter(id=current_method_id).exists():
                return Response(
                    {
                        "error": "Указанный метод не принадлежит данной странице исследований"
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )

            current_method = ResearchMethod.objects.get(id=current_method_id)
            serializer = ResearchMethodSerializer(current_method)

            return Response({"current_method": serializer.data})

        except ResearchMethod.DoesNotExist:
            return Response(
                {"error": "Метод исследования не найден"},
                status=status.HTTP_404_NOT_FOUND,
            )
        except Exception as e:
            return Response(
                {"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    def create(self, request):
        """
        Создает страницу расчетов для выбранного типа исследования.
        Может быть привязана к подразделению или напрямую к лаборатории.
        """
        try:
            serializer = self.get_serializer(data=request.data)
            serializer.is_valid(raise_exception=True)

            research_type = serializer.validated_data["type"]
            laboratory_id = serializer.validated_data["laboratory"].id
            department = serializer.validated_data.get("department")
            department_id = department.id if department else None

            # Проверяем существование страницы
            existing_page = ResearchObject.objects.filter(
                laboratory_id=laboratory_id,
                department_id=department_id,
                type=research_type,
                is_deleted=False,
            ).first()

            if existing_page:
                return Response(
                    {"message": "Страница расчетов уже существует"},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            # Если есть department_id, проверяем его
            if department_id:
                if department.laboratory_id != laboratory_id:
                    return Response(
                        {
                            "message": "Подразделение не принадлежит указанной лаборатории"
                        },
                        status=status.HTTP_400_BAD_REQUEST,
                    )

            research_object = serializer.save()

            return Response(serializer.data, status=status.HTTP_201_CREATED)

        except Exception as e:
            return Response(
                {"message": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    def update(self, request, *args, **kwargs):
        instance = self.get_object()

        # Если в запросе есть research_methods, обрабатываем их отдельно
        if "research_methods" in request.data:
            method_ids = request.data.pop("research_methods")
            request.data["research_method_ids"] = method_ids

        serializer = self.get_serializer(instance, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)

        return Response(serializer.data)

    def list(self, request):
        """
        Возвращает список страниц расчетов с фильтрацией по лаборатории и подразделению.
        """
        laboratory_id = request.query_params.get("laboratory_id")
        department_id = request.query_params.get("department_id")

        queryset = self.get_queryset()

        if laboratory_id:
            queryset = queryset.filter(laboratory_id=laboratory_id)
        if department_id:
            queryset = queryset.filter(department_id=department_id)

        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        instance.is_deleted = True
        instance.deleted_at = timezone.now()
        instance.save()
        return Response(status=status.HTTP_204_NO_CONTENT)


def _round_to_significant_figures(number, significant_figures):
    if number == 0:
        return 0

    d = Decimal(str(float(number)))
    str_num = f"{d:E}"
    mantissa, exp = str_num.split("E")
    exp = int(exp)
    mantissa = mantissa.replace(".", "").rstrip("0")

    if len(mantissa) > significant_figures:
        mantissa = str(round(int(mantissa[: significant_figures + 1]) / 10))

    mantissa = mantissa.ljust(significant_figures, "0")

    if exp >= 0:
        if exp + 1 >= len(mantissa):
            result = Decimal(mantissa + "0" * (exp + 1 - len(mantissa)))
        else:
            result = Decimal(mantissa[: exp + 1] + "." + mantissa[exp + 1 :])
    else:
        result = Decimal("0." + "0" * (-exp - 1) + mantissa)

    return result


def _get_decimal_places(number):
    str_num = str(Decimal(str(float(number))))
    if "." not in str_num:
        return 0
    return len(str_num.split(".")[1])


def round_result(result, rounding_type, rounding_decimal):
    if rounding_type == "decimal":
        d = Decimal(str(float(result)))
        # Используем ROUND_HALF_UP для округления 0.5 вверх
        return d.quantize(Decimal("0.1") ** rounding_decimal, rounding=ROUND_HALF_UP)
    else:
        return _round_to_significant_figures(result, rounding_decimal)


def _replace_subscript_digits(text):
    subscript_map = {
        "₀": "0",
        "₁": "1",
        "₂": "2",
        "₃": "3",
        "₄": "4",
        "₅": "5",
        "₆": "6",
        "₇": "7",
        "₈": "8",
        "₉": "9",
        "ₐ": "a",
        "ₑ": "e",
        "ₕ": "h",
        "ᵢ": "i",
        "ⱼ": "j",
        "ₖ": "k",
        "ₗ": "l",
        "ₘ": "m",
        "ₙ": "n",
        "ₒ": "o",
        "ₚ": "p",
        "ᵣ": "r",
        "ₛ": "s",
        "ₜ": "t",
        "ᵤ": "u",
        "ᵥ": "v",
        "ₓ": "x",
        "ᵦ": "β",
        "ᵧ": "γ",
        "ᵨ": "ρ",
        "ᵩ": "φ",
        "ᵪ": "χ",
    }

    result = text
    for subscript, normal in subscript_map.items():
        result = result.replace(subscript, normal)
    return result


def evaluate_formula(formula, variables, is_condition=False):
    """
    Вычисляет результат формулы с использованием безопасного eval.

    Args:
        formula (str): Формула для вычисления
        variables (dict): Словарь переменных и их значений
        is_condition (bool): Флаг, указывающий что это вычисление условия

    Returns:
        Decimal: Результат вычисления или bool для условий
    """
    # Создаем копию переменных с замененными индексами
    decimal_vars = {}
    for name, value in variables.items():
        try:
            if isinstance(value, str):
                value = value.strip().replace(",", ".")
            new_name = _replace_subscript_digits(name)
            decimal_vars[new_name] = float(value)
        except Exception as e:
            raise ValueError(
                f"Ошибка преобразования значения {name} = {value} в число: {str(e)}"
            )

    try:
        formula = _replace_subscript_digits(formula)
        formula = formula.replace("×", "*").replace("÷", "/")

        # Создаем безопасный контекст для вычислений
        safe_dict = {"abs": abs, "pow": pow, "round": round, "max": max, "min": min}
        safe_dict.update(decimal_vars)

        if is_condition:
            # Для условий разбиваем формулу на части
            for operator in ["<=", ">=", ">", "<", "="]:
                if operator in formula:
                    left, right = formula.split(operator)
                    # Вычисляем левую и правую части
                    left_result = float(eval(left, {"__builtins__": {}}, safe_dict))
                    right_result = float(eval(right, {"__builtins__": {}}, safe_dict))

                    # Добавляем небольшую погрешность для сравнения чисел с плавающей точкой
                    epsilon = 1e-10

                    if operator == "<=":
                        return left_result <= (right_result + epsilon)
                    elif operator == ">=":
                        return left_result >= (right_result - epsilon)
                    elif operator == ">":
                        return left_result > (right_result + epsilon)
                    elif operator == "<":
                        return left_result < (right_result - epsilon)
                    else:  # =
                        return abs(left_result - right_result) < 1e-10

            raise ValueError(
                f"Неподдерживаемый оператор сравнения в формуле: {formula}"
            )
        else:
            # Для обычных формул
            result = float(eval(formula, {"__builtins__": {}}, safe_dict))
            return Decimal(str(result))

    except Exception as e:
        raise ValueError(f"Ошибка при вычислении формулы '{formula}': {str(e)}")


@api_view(["POST"])
@permission_classes([AllowAny])
def calculate_result(request):
    try:
        logger.info("Начало расчета")

        input_data = request.data.get("input_data", {})
        research_method = request.data.get("research_method", {})

        logger.info(f"Полученные входные данные: {input_data}")
        logger.info(
            f"Метод исследования: {research_method['name']} (ID: {research_method['id']})"
        )
        logger.info(f"Формула расчета: {research_method['formula']}")
        logger.info(f"Погрешность: {research_method['measurement_error']}")

        if not input_data or not research_method:
            logger.error("Отсутствуют необходимые данные для расчета")
            return Response(
                {
                    "error": "Необходимо предоставить входные данные и метод исследования"
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        logger.info("Начало проверки условий сходимости")
        satisfied_conditions = []

        for condition in research_method["convergence_conditions"]["formulas"]:
            try:
                logger.info(f"Проверка условия: {condition['formula']}")
                condition_result = evaluate_formula(
                    condition["formula"], input_data, is_condition=True
                )
                logger.info(
                    f"Результат проверки условия: {condition_result} (тип: {condition['convergence_value']})"
                )

                if condition_result:
                    satisfied_conditions.append(condition["convergence_value"])
                    logger.info(
                        f"Условие {condition['formula']} выполнено, тип: {condition['convergence_value']}"
                    )
            except Exception as e:
                logger.error(f"Ошибка при проверке условия сходимости: {str(e)}")
                return Response(
                    {"error": f"Ошибка при проверке условия сходимости: {str(e)}"},
                    status=status.HTTP_400_BAD_REQUEST,
                )

        # Определяем итоговый результат сходимости
        logger.info(f"Все выполненные условия: {satisfied_conditions}")

        # Проверяем наличие особых условий
        special_conditions = [
            cond
            for cond in satisfied_conditions
            if cond in ["traces", "unsatisfactory", "absence"]
        ]

        if special_conditions:
            # Если есть особые условия, возвращаем их через запятую
            convergence_result = ", ".join(special_conditions)
            logger.info(f"Найдены особые условия: {convergence_result}")

            return Response(
                {
                    "convergence": convergence_result,
                    "intermediate_results": {},
                    "result": None,
                    "measurement_error": None,
                    "unit": research_method["unit"],
                },
                status=status.HTTP_200_OK,
            )

        # Если особых условий нет, считаем результат удовлетворительным
        convergence_result = "satisfactory"
        logger.info("Условия сходимости удовлетворительны, продолжаем расчет")

        # Если сходимость удовлетворительная, вычисляем результат
        result = None
        measurement_error = None
        intermediate_results = {}
        variables = dict(input_data)

        if convergence_result == "satisfactory":
            logger.info("Начало вычисления промежуточных результатов")
            # Сначала вычисляем промежуточные результаты
            for field in research_method["intermediate_data"]["fields"]:
                # Пропускаем поля с пустыми именами или формулами
                if not field["name"].strip() or not field["formula"].strip():
                    logger.info(f"Пропущено пустое промежуточное поле: {field}")
                    continue

                try:
                    logger.info(
                        f"Вычисление промежуточного результата: {field['name']}, формула: {field['formula']}"
                    )
                    intermediate_value = evaluate_formula(field["formula"], variables)
                    logger.info(
                        f"Промежуточный результат {field['name']} = {intermediate_value}"
                    )
                    intermediate_results[field["name"]] = str(intermediate_value)
                    variables[field["name"]] = intermediate_value
                except Exception as e:
                    logger.error(
                        f"Ошибка при вычислении промежуточного результата {field['name']}: {str(e)}"
                    )
                    return Response(
                        {
                            "error": f"Ошибка при вычислении промежуточного результата: {str(e)}"
                        },
                        status=status.HTTP_400_BAD_REQUEST,
                    )

            # Затем вычисляем основной результат
            try:
                logger.info(
                    f"Вычисление основного результата по формуле: {research_method['formula']}"
                )
                logger.info(f"Используемые переменные: {variables}")
                result = evaluate_formula(research_method["formula"], variables)
                logger.info(f"Неокругленный результат: {result}")
            except Exception as e:
                logger.error(f"Ошибка при вычислении основного результата: {str(e)}")
                return Response(
                    {"error": f"Ошибка при вычислении результата: {str(e)}"},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            # Округляем результат согласно настройкам
            logger.info(
                f"Округление результата: тип={research_method['rounding_type']}, знаков={research_method['rounding_decimal']}"
            )
            result = round_result(
                result,
                research_method["rounding_type"],
                research_method["rounding_decimal"],
            )
            logger.info(f"Окончательный результат после округления: {result}")

            # Вычисляем количество знаков после запятой в результате
            result_decimal_places = (
                len(str(result).split(".")[-1]) if "." in str(result) else 0
            )
            logger.info(
                f"Количество знаков после запятой в результате: {result_decimal_places}"
            )

            # Вычисляем погрешность
            try:
                error_config = research_method["measurement_error"]
                logger.info(f"Вычисление погрешности: {error_config}")

                if error_config["type"] == "fixed":
                    measurement_error = float(error_config["value"])
                elif error_config["type"] == "formula":
                    variables["result"] = result
                    measurement_error = float(
                        evaluate_formula(error_config["value"], variables)
                    )
                elif error_config["type"] == "range":
                    for range_item in error_config["ranges"]:
                        if evaluate_formula(
                            range_item["formula"], variables, is_condition=True
                        ):
                            measurement_error = float(range_item["value"])
                            break
                    if measurement_error is None:
                        logger.warning("Не найден подходящий диапазон для погрешности")
                        measurement_error = 0

                # Округляем погрешность до того же количества знаков после запятой, что и результат
                if measurement_error is not None:
                    measurement_error = round(
                        Decimal(str(measurement_error)), result_decimal_places
                    )
                    logger.info(f"Погрешность после округления: {measurement_error}")

            except Exception as e:
                logger.error(f"Ошибка при вычислении погрешности: {str(e)}")
                return Response(
                    {"error": f"Ошибка при вычислении погрешности: {str(e)}"},
                    status=status.HTTP_400_BAD_REQUEST,
                )

        response_data = {
            "convergence": convergence_result,
            "intermediate_results": intermediate_results,
            "result": str(result) if result is not None else None,
            "measurement_error": (
                str(measurement_error) if measurement_error is not None else None
            ),
            "unit": research_method["unit"],
        }
        logger.info(f"Подготовлен ответ: {response_data}")

        return Response(response_data, status=status.HTTP_200_OK)

    except Exception as e:
        logger.error(f"Необработанная ошибка при расчете: {str(e)}", exc_info=True)
        return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
